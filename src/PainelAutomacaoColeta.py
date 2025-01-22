import os
import sys
import time
import shutil
import re
from PyQt6.QtWidgets import (
    QLabel, QComboBox, QWidget, QGridLayout, QMessageBox,
    QTextEdit, QFileDialog, QLineEdit, QVBoxLayout, QPushButton, QApplication
)
from PyQt6.QtCore import QThreadPool, Qt, QSettings, QRunnable, QMutex
from PyQt6.QtGui import QTextCursor
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from PyPDF2 import PdfReader

# Configurações de estilo
from GerenEstilos import (
    estilo_combo_box_light, estilo_hover,
    estilo_log_light, estilo_label_light, campo_qline_light, 
    estilo_log_dark, campo_qline_dark, estilo_combo_box_dark,
    estilo_label_dark
)

###############################################
# PARTE 1: LÓGICA DE PROCESSAMENTO (BACKEND)
###############################################

class TarefaAutomacao(QRunnable):
    """Classe que executa a automação em uma thread separada"""
    
    def __init__(self, automator, dados_usuario, funcao_log):
        super().__init__()
        self.automator = automator
        self.dados_usuario = dados_usuario
        self.funcao_log = funcao_log

    def run(self):
        """Método principal que inicia o processo de automação"""
        try:
            self.funcao_log("Iniciando automação...", area="tecnico")
            if hasattr(self.automator, 'flag_parar') and self.automator.flag_parar:
                self.funcao_log("Automação cancelada antes de iniciar", area="tecnico")
                return
            
            self.automator.executar_automacao(self.dados_usuario)
            self.funcao_log("Automação concluída!", area="tecnico")
            
        except Exception as erro:
            self.funcao_log(f"Erro durante a automação: {str(erro)}", area="tecnico")

class PararAutomacao:
    """Classe responsável por interromper a automação em execução"""
    
    def __init__(self, automator):
        self.automator = automator

    def parar(self):
        """Para a automação e fecha todos os navegadores"""
        if hasattr(self.automator, 'flag_parar'):
            self.automator.flag_parar = True
            self.automator.parent.log_mensagem("Parando automação...", area="tecnico", cor="red")
            self.fechar_navegadores()
        else:
            self.automator.parent.log_mensagem("Nenhuma automação em execução", area="tecnico")

    def fechar_navegadores(self):
        """Fecha todas as instâncias do navegador abertas"""
        if hasattr(self.automator, 'drivers') and self.automator.drivers:
            self.automator.parent.log_mensagem("Fechando navegadores...", area="tecnico")
            for driver in self.automator.drivers:
                try:
                    driver.quit()
                except Exception as erro:
                    self.automator.parent.log_mensagem(f"Erro ao fechar navegador: {erro}", area="tecnico")
            self.automator.drivers.clear()

class Blume:
    """Classe principal que implementa a automação para a operadora Blume"""
    
    def __init__(self, parent, caminho_dados):
        self.parent = parent
        self.caminho_dados = caminho_dados
        self.planilha = load_workbook(caminho_dados).active
        self.mutex = QMutex()
        self.flag_parar = False
        self.drivers = []

    def inicializar_navegador(self):
        """Configura e inicia uma nova instância do navegador Chrome"""
        try:
            opcoes = webdriver.ChromeOptions()
            opcoes.add_argument("--disable-extensions")
            opcoes.add_argument("--disable-popup-blocking")
            opcoes.add_argument("--headless")

            self.parent.log_mensagem("Abrindo navegador...", area="tecnico")
            driver = webdriver.Chrome(
                service=Service(ChromeDriverManager().install()),
                options=opcoes
            )
            self.drivers.append(driver)
            return driver
            
        except Exception as erro:
            self.parent.log_mensagem(f"Falha ao iniciar navegador: {erro}", area="tecnico")
            raise

    def executar_automacao(self, dados_usuario):
        """Fluxo principal de execução da automação"""
        self.parent.log_mensagem("Iniciando coleta para Blume...", area="tecnico")

        if self.verificar_coleta_completa():
            self.parent.log_mensagem("Todas faturas já foram coletadas!", area="tecnico")
            return

        for usuario in dados_usuario:
            if self.flag_parar:
                self.parent.log_mensagem("Processo interrompido!", area="tecnico")
                self.fechar_navegadores()
                return

            if usuario['STATUS'] == 'COLETADO IA':
                continue

            driver = None
            try:
                driver = self.inicializar_navegador()
                wait = WebDriverWait(driver, 15)
                driver.get("https://portal.blumetelecom.com.br")
                self.fazer_login(driver, wait, usuario)
                self.processar_boletos(driver, wait, usuario)
                
            except Exception as erro:
                self.parent.log_mensagem(f"Erro no processamento: {str(erro)}", area="tecnico")
            finally:
                if driver:
                    driver.quit()
                    self.drivers.remove(driver)

    def fazer_login(self, driver, wait, dados_usuario):
        """Realiza o login no portal da Blume"""
        try:
            self.parent.log_mensagem(f"Realizando login no acesso: {dados_usuario['LOGIN']}", area="tecnico")
            wait.until(lambda driver: driver.execute_script("return document.readyState") == "complete")

            campo_email = wait.until(EC.presence_of_element_located((By.ID, "loginUsername")))
            driver.execute_script("arguments[0].scrollIntoView(true);", campo_email)
            campo_email.send_keys(dados_usuario['LOGIN'])

            campo_senha = wait.until(EC.element_to_be_clickable((By.NAME, "password")))
            driver.execute_script("arguments[0].scrollIntoView(true);", campo_senha)
            campo_senha.clear()
            campo_senha.send_keys(dados_usuario['SENHA'])

            time.sleep(1)
            botao_login = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "MuiButton-label")))
            driver.execute_script("arguments[0].scrollIntoView(true);", botao_login)
            botao_login.click()

            time.sleep(2)
            driver.get("https://portal.blumetelecom.com.br/billings")

            self.parent.log_mensagem("Login realizado com sucesso!", area="tecnico")

        except Exception as erro:
            self.parent.log_mensagem(f"Falha no login: {erro}", area="tecnico")
            raise

    def processar_boletos(self, driver, wait, dados_usuario):
        """Processa os boletos disponíveis no portal"""
        try:
            ids_processados = set()

            while True:
                driver.get("https://portal.blumetelecom.com.br/billings")

                try:
                    mensagem_sem_faturas = wait.until(
                        EC.presence_of_element_located(
                            (By.XPATH, "//h5[contains(text(), 'Você não possui faturas em aberto')]"))
                    )
                    self.parent.log_mensagem("Nenhuma fatura disponível", area="tecnico")
                    self.atualizar_status_planilha(dados_usuario['LOGIN'], 'INDISPONIVEL')
                    return
                except:
                    pass

                try:
                    botoes_boleto = wait.until(
                        EC.presence_of_all_elements_located((By.XPATH, "//span[text()='Pagar boleto']"))
                    )
                except Exception as erro:
                    self.parent.log_mensagem(f"Erro ao buscar boletos: {erro}", area="tecnico")
                    break

                for indice, botao in enumerate(botoes_boleto):
                    id_boleto = f"botao-{indice}"
                    if id_boleto in ids_processados:
                        continue

                    try:
                        wait.until(EC.element_to_be_clickable(botao))
                        driver.execute_script("arguments[0].scrollIntoView(true);", botao)
                        botao.click()
                        self.baixar_boleto(wait, dados_usuario, indice + 1)
                        ids_processados.add(id_boleto)
                    except Exception as erro:
                        self.parent.log_mensagem(f"Erro ao processar boleto {indice + 1}: {erro}", area="tecnico")

                if len(ids_processados) >= len(botoes_boleto):
                    break

        except Exception as erro:
            self.parent.log_mensagem(f"Erro geral no processamento: {erro}", area="tecnico")
            raise

    def baixar_boleto(self, wait, dados_usuario, indice):
        """Realiza o download e processamento do boleto"""
        try:
            botao_download = wait.until(
                EC.presence_of_element_located((By.XPATH, "//p[text()='Baixar boleto']"))
            )
            botao_download.click()

            caminho_download = os.path.join(os.path.expanduser('~'), 'Downloads')
            arquivo = self.aguardar_download(caminho_download)

            if arquivo:
                contrato = self.extrair_contrato_pdf(arquivo)
                if contrato:
                    self.processar_arquivo_baixado(contrato, arquivo, dados_usuario)
        except Exception as erro:
            self.parent.log_mensagem(f"Erro no download: {erro}", area="tecnico")

    def aguardar_download(self, pasta):
        """Aguarda a conclusão do download do arquivo"""
        tempo_limite = 60
        while tempo_limite > 0:
            arquivos = [f for f in os.listdir(pasta) if f.endswith(".pdf")]
            if arquivos:
                arquivo_recente = os.path.join(pasta, max(arquivos, key=lambda f: os.path.getmtime(os.path.join(pasta, f))))
                if not arquivo_recente.endswith(".crdownload"):
                    return arquivo_recente
            time.sleep(1)
            tempo_limite -= 1
        return None

    def extrair_contrato_pdf(self, caminho_pdf):
        """Extrai o número do contrato de um arquivo PDF"""
        try:
            leitor = PdfReader(caminho_pdf)
            texto = "".join([pagina.extract_text() for pagina in leitor.pages])
            contrato = re.search(r'\b0{3,}(\d{3,})\b', texto)
            return contrato.group(1) if contrato else None
        except Exception as erro:
            self.parent.log_mensagem(f"Erro na leitura do PDF: {erro}", area="tecnico")
            return None

    def processar_arquivo_baixado(self, contrato, arquivo, dados_usuario):
        """Processa o arquivo baixado e atualiza a planilha"""
        contrato = contrato.lstrip('0')
        if self.verificar_contrato_planilha(contrato):
            nomenclatura = self.obter_nomenclatura(contrato)
            destino = os.path.join(self.parent.pasta_salvamento, f"{nomenclatura}.pdf")
            shutil.move(arquivo, destino)
            self.parent.log_mensagem(nomenclatura, area="faturas")
            self.atualizar_status_planilha(contrato, 'COLETADO IA')
        else:
            self.mover_arquivo_nao_encontrado(arquivo)

    def mover_arquivo_nao_encontrado(self, arquivo):
        """Move arquivos não identificados para pasta específica"""
        pasta_erro = os.path.join(self.parent.pasta_salvamento, "Boleto não encontrado")
        os.makedirs(pasta_erro, exist_ok=True)
        shutil.move(arquivo, os.path.join(pasta_erro, os.path.basename(arquivo)))

    def verificar_contrato_planilha(self, contrato):
        """Verifica se o contrato existe na planilha"""
        for linha in self.planilha.iter_rows(min_row=2, values_only=True):
            if str(linha[4]).lstrip('0') == contrato:
                return True
        return False

    def obter_nomenclatura(self, contrato):
        """Obtém a nomenclatura correta do contrato"""
        for linha in self.planilha.iter_rows(min_row=2, values_only=True):
            if str(linha[4]).lstrip('0') == contrato:
                return linha[12]
        return None

    def atualizar_status_planilha(self, identificador, status):
        """Atualiza o status na planilha Excel"""
        self.mutex.lock()
        try:
            for linha in self.planilha.iter_rows(min_row=2):
                if str(linha[8].value) == identificador or str(linha[4].value).lstrip('0') == identificador:
                    linha[11].value = status
                    self.planilha.parent.save(self.caminho_dados)
                    break
        finally:
            self.mutex.unlock()

    def verificar_coleta_completa(self):
        """Verifica se todas as faturas foram coletadas"""
        for linha in self.planilha.iter_rows(min_row=2, values_only=True):
            if linha[11] != 'COLETADO IA':
                return False
        return True

    def fechar_navegadores(self):
        """Fecha todas as instâncias do navegador"""
        for driver in self.drivers:
            try:
                driver.quit()
            except:
                pass
        self.drivers.clear()

###############################################
# PARTE 2: INTERFACE DO USUÁRIO (FRONTEND)
###############################################

class InterfaceAutoBlume(QWidget):
    """Classe que constrói e gerencia a interface gráfica"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.planilha = None
        self.caminho_dados = ""
        self.pasta_salvamento = ""
        self.threads = QThreadPool()
        self.inicializar_interface()
        self.carregar_configuracoes()

    def apply_styles(self, dark_mode):
        """Aplica estilos dinamicamente com base no tema"""
        estilo_combo = estilo_combo_box_dark() if dark_mode else estilo_combo_box_light()
        estilo_campo = campo_qline_dark() if dark_mode else campo_qline_light()
        estilo_log = estilo_log_dark() if dark_mode else estilo_log_light()
        estilo_label = estilo_label_dark() if dark_mode else estilo_label_light()

        for btn in [self.botao_pasta, self.botao_planilha, self.botao_iniciar]:
            estilo_hover(btn, dark_mode)

        # Aplica estilos nos componentes
        self.combo_operadora.setStyleSheet(estilo_combo)
        self.campo_pasta.setStyleSheet(estilo_campo)
        self.campo_planilha.setStyleSheet(estilo_campo)
        self.log_tecnico.setStyleSheet(estilo_log)
        self.log_faturas.setStyleSheet(estilo_log)
        self.rotulo_pasta.setStyleSheet(estilo_label)
        self.rotulo_planilha.setStyleSheet(estilo_label)
        self.rotulo_operadora.setStyleSheet(estilo_label)

    def inicializar_interface(self):
        """Configura todos os elementos da interface gráfica"""
        self.setWindowTitle("Automação Blume")

        # Layout principal
        layout_principal = QVBoxLayout(self)
        layout_principal.setContentsMargins(10, 10, 10, 10)
        layout_principal.setSpacing(10)

        # Seção de configurações
        layout_superior = QGridLayout()
        
        # Campo de pasta de salvamento
        self.rotulo_pasta = QLabel("Pasta de Salvamento:")
        self.rotulo_pasta.setStyleSheet(estilo_label_light())
        self.campo_pasta = QLineEdit()
        self.campo_pasta.setReadOnly(True)
        self.campo_pasta.setStyleSheet(campo_qline_light())
        self.botao_pasta = QPushButton("Selecionar Pasta")
        estilo_hover(self.botao_pasta)
        self.botao_pasta.clicked.connect(self.selecionar_pasta)

        # Campo de planilha de dados
        self.rotulo_planilha = QLabel("Planilha de Dados:")
        self.rotulo_planilha.setStyleSheet(estilo_label_light())
        self.campo_planilha = QLineEdit()
        self.campo_planilha.setReadOnly(True)
        self.campo_planilha.setStyleSheet(campo_qline_light())
        self.botao_planilha = QPushButton("Selecionar Arquivo")
        estilo_hover(self.botao_planilha)
        self.botao_planilha.clicked.connect(self.selecionar_planilha)

        # Seleção de operadora
        self.rotulo_operadora = QLabel("Operadora:")
        self.rotulo_operadora.setStyleSheet(estilo_label_light())
        self.combo_operadora = QComboBox()
        self.combo_operadora.addItem("Selecione uma planilha primeiro")
        self.combo_operadora.setStyleSheet(estilo_combo_box_light())

        # Botão de controle
        self.botao_iniciar = QPushButton("Iniciar Automação")
        estilo_hover(self.botao_iniciar)
        self.botao_iniciar.clicked.connect(self.alternar_automacao)

        # Adicionando elementos ao layout
        layout_superior.addWidget(self.rotulo_pasta, 0, 0)
        layout_superior.addWidget(self.campo_pasta, 0, 1)
        layout_superior.addWidget(self.botao_pasta, 0, 2)
        
        layout_superior.addWidget(self.rotulo_planilha, 1, 0)
        layout_superior.addWidget(self.campo_planilha, 1, 1)
        layout_superior.addWidget(self.botao_planilha, 1, 2)
        
        layout_superior.addWidget(self.rotulo_operadora, 2, 0)
        layout_superior.addWidget(self.combo_operadora, 2, 1)
        layout_superior.addWidget(self.botao_iniciar, 2, 2)

        # Área de logs
        self.log_tecnico = QTextEdit()
        self.log_tecnico.setPlaceholderText("Logs técnicos...")
        self.log_tecnico.setReadOnly(True)
        self.log_tecnico.setStyleSheet(estilo_log_light())
        
        self.log_faturas = QTextEdit()
        self.log_faturas.setPlaceholderText("Faturas coletadas...")
        self.log_faturas.setReadOnly(True)
        self.log_faturas.setStyleSheet(estilo_log_light())

        # Montagem final
        layout_principal.addLayout(layout_superior)
        layout_principal.addWidget(self.log_tecnico)
        layout_principal.addWidget(self.log_faturas)

    def carregar_configuracoes(self):
        """Carrega as configurações salvas anteriormente"""
        config = QSettings("config.ini", QSettings.Format.IniFormat)
        self.pasta_salvamento = config.value("pasta_salvamento", "")
        self.caminho_dados = config.value("caminho_dados", "")
        self.campo_pasta.setText(self.pasta_salvamento)

    def salvar_configuracoes(self):
        """Salva as configurações atuais"""
        config = QSettings("config.ini", QSettings.Format.IniFormat)
        config.setValue("pasta_salvamento", self.pasta_salvamento)
        config.setValue("caminho_dados", self.caminho_dados)

    def selecionar_pasta(self):
        """Seleciona a pasta para salvar os boletos"""
        pasta = QFileDialog.getExistingDirectory(self, "Selecionar Pasta de Salvamento")
        if pasta:
            self.pasta_salvamento = pasta
            self.campo_pasta.setText(pasta)
            self.salvar_configuracoes()

    def selecionar_planilha(self):
        """Seleciona o arquivo Excel com os dados"""
        arquivo, _ = QFileDialog.getOpenFileName(
            self,
            "Selecionar Planilha",
            "",
            "Arquivos Excel (*.xlsx *.xlsm)"
        )
        if arquivo:
            self.carregar_planilha(arquivo)

    def carregar_planilha(self, caminho):
        """Carrega e processa o arquivo Excel selecionado"""
        try:
            self.planilha = load_workbook(caminho).active
            self.caminho_dados = caminho
            self.campo_planilha.setText(caminho)
            self.salvar_configuracoes()

            operadoras = set()
            for linha in self.planilha.iter_rows(min_row=2, values_only=True):
                if linha[3]:
                    operadoras.add(linha[3])
            
            self.combo_operadora.clear()
            self.combo_operadora.addItems(operadoras)
            
        except Exception as erro:
            self.log_mensagem(f"Erro ao carregar planilha: {erro}", area="tecnico")
            QMessageBox.critical(self, "Erro", f"Falha ao carregar arquivo: {erro}")

    def alternar_automacao(self):
        """Controla o início/parada da automação"""
        if self.botao_iniciar.text() == "Iniciar Automação":
            self.iniciar_automacao()
            self.botao_iniciar.setText("Parar Automação")
        else:
            self.parar_automacao()
            self.botao_iniciar.setText("Iniciar Automação")

    def iniciar_automacao(self):
        """Inicia o processo de automação"""
        if not self.validar_campos():
            return

        operadora = self.combo_operadora.currentText()
        dados = self.obter_dados_usuario(operadora)

        try:
            self.automator = Blume(self, self.caminho_dados)
            tarefa = TarefaAutomacao(self.automator, dados, self.log_mensagem)
            self.threads.start(tarefa)
        except Exception as erro:
            self.log_mensagem(f"Falha ao iniciar: {erro}", area="tecnico")

    def parar_automacao(self):
        """Interrompe a automação em execução"""
        if hasattr(self, 'automator'):
            PararAutomacao(self.automator).parar()
        self.botao_iniciar.setText("Iniciar Automação")

    def validar_campos(self):
        """Valida os campos antes de iniciar"""
        erros = []
        if not self.pasta_salvamento:
            erros.append("Selecione uma pasta de salvamento")
        if not self.caminho_dados:
            erros.append("Selecione uma planilha de dados")
        if self.combo_operadora.currentText() == "Selecione uma planilha primeiro":
            erros.append("Selecione uma operadora válida")
        
        if erros:
            QMessageBox.warning(self, "Atenção", "\n".join(erros))
            return False
        return True

    def obter_dados_usuario(self, operadora):
        """Extrai os dados da planilha para a operadora selecionada"""
        dados = []
        for linha in self.planilha.iter_rows(min_row=2, values_only=True):
            if linha[3] == operadora and linha[11] != 'COLETADO IA':
                dados.append({
                    "FORNECEDOR": linha[0],
                    "REFERÊNCIA": linha[1],
                    "CLIENTE": linha[2],
                    "OPERADORA": linha[3],
                    "IDENTIFICAÇÃO": linha[4],
                    "CÓDIGO": linha[5],
                    "PA": linha[6],
                    "INDENTIFICAÇÃO INTERNA": linha[7],
                    "LOGIN": linha[8],
                    "SENHA": linha[9],
                    "VENCIMENTO": linha[10],
                    "STATUS": linha[11],
                    "NOMENCLATURA": linha[12]
                })
        return dados

    def log_mensagem(self, mensagem, area="tecnico", cor=None):
        """Exibe mensagens nas áreas de log correspondentes"""
        area_log = self.log_tecnico if area == "tecnico" else self.log_faturas

        if cor:
            mensagem = f'<span style="color: {cor};">{mensagem}</span>'
        area_log.append(mensagem)
        cursor = area_log.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)
        area_log.setTextCursor(cursor)