import os
import sys
from PyQt6.QtWidgets import (
    QLabel, QComboBox, QWidget, QGridLayout, QMessageBox, QTextEdit, QFileDialog, QLineEdit, QVBoxLayout, QPushButton
)
from PyQt6.QtCore import QThreadPool, Qt, QSettings
from PyQt6.QtGui import QTextCursor
from openpyxl import load_workbook

#Backend
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'backend')))
from AutoBlume import Blume, AutomationTask, StopAutomation

#Estilos e gerenciamento
from GerenEstilos import (
    estilo_label_light, estilo_combo_box_light, estilo_hover,
    campo_qline_light, estilo_log_light
)


class GuiAutoBlume(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.workbook = None
        self.sheet = None
        self.data_path = ""
        self.save_directory = ""
        self.threadpool = QThreadPool()
        self.stopflag = False
        self.init_ui()

    def init_ui(self):
        """Inicializa a interface gráfica."""
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(10, 10, 10, 10)
        self.layout.setSpacing(10)

        top_layout = QGridLayout()

        # Aplicar estilo ao QLabel
        self.label_salvamento = QLabel("Local de Salvamento:", self)
        self.label_salvamento.setStyleSheet(estilo_label_light())
        self.label_planilha = QLabel("Planilha de dados:", self)
        self.label_planilha.setStyleSheet(estilo_label_light())
        self.label_operadora = QLabel("Selecionar Operadora:", self)
        self.label_operadora.setStyleSheet(estilo_label_light())

        top_layout.addWidget(self.label_salvamento, 0, 0, alignment=Qt.AlignmentFlag.AlignLeft)
        top_layout.addWidget(self.label_planilha, 1, 0, alignment=Qt.AlignmentFlag.AlignLeft)
        top_layout.addWidget(self.label_operadora, 2, 0, alignment=Qt.AlignmentFlag.AlignLeft)

        # Diretório de Salvamento
        self.save_dir_field = QLineEdit(self)
        self.save_dir_field.setReadOnly(True)
        self.save_dir_field.setText(self.save_directory)
        self.save_dir_field.setStyleSheet(campo_qline_light())
        top_layout.addWidget(self.save_dir_field, 0, 1)

        self.save_dir_button = QPushButton("Selecionar Pasta", self)
        estilo_hover(self.save_dir_button)
        self.save_dir_button.clicked.connect(self.select_save_directory)
        top_layout.addWidget(self.save_dir_button, 0, 2)

        # Seleção de Planilha
        self.planilha_field = QLineEdit(self)
        self.planilha_field.setReadOnly(True)
        self.planilha_field.setText(self.data_path)
        self.planilha_field.setStyleSheet(campo_qline_light())
        top_layout.addWidget(self.planilha_field, 1, 1)

        self.planilha_button = QPushButton("Selecionar Planilha", self)
        estilo_hover(self.planilha_button)
        self.planilha_button.clicked.connect(self.select_data_file)
        top_layout.addWidget(self.planilha_button, 1, 2)

        # Seleção de Operadora
        self.operadora_combo = QComboBox(self)
        self.operadora_combo.addItem("Selecione uma planilha primeiro")
        self.operadora_combo.setStyleSheet(estilo_combo_box_light())
        top_layout.addWidget(self.operadora_combo, 2, 1)

        # Botão de Iniciar/Parar Automação
        self.confirm_button = QPushButton("Iniciar automação", self)
        estilo_hover(self.confirm_button)
        self.confirm_button.clicked.connect(self.toggle_automation)  # Alterado para toggle_automation
        top_layout.addWidget(self.confirm_button, 2, 2)

        self.layout.addLayout(top_layout)

        # Layout dos Logs
        logs_layout = QGridLayout()

        self.log_tecnico_area = QTextEdit(self)
        self.log_tecnico_area.setReadOnly(True)
        self.log_tecnico_area.setPlaceholderText("Log técnico")
        self.log_tecnico_area.setStyleSheet(estilo_log_light())
        logs_layout.addWidget(self.log_tecnico_area, 0, 0, 2, 2)

        self.faturas_coletadas_area = QTextEdit(self)
        self.faturas_coletadas_area.setReadOnly(True)
        self.faturas_coletadas_area.setPlaceholderText("Log faturas coletadas")
        self.faturas_coletadas_area.setStyleSheet(estilo_log_light())
        logs_layout.addWidget(self.faturas_coletadas_area, 0, 2, 2, 2)

        self.layout.addLayout(logs_layout)

    def select_data_file(self):
        last_dir = QSettings(self.parent.settings_path, QSettings.Format.IniFormat).value("last_open_dir", "")
        file_path, _ = QFileDialog.getOpenFileName(self, "Selecionar Planilha de Dados", last_dir, "Arquivos Excel (*.xlsx *.xlsm)")

        if file_path:
            if not os.path.isfile(file_path):
                QMessageBox.critical(self, "Erro", "O arquivo selecionado não é válido.")
            return
        self.load_data_file(file_path)

    def load_data_file(self, file_path):
        """Carrega a planilha de dados e atualiza a interface."""
        try:
            self.workbook = load_workbook(file_path)
            self.sheet = self.workbook.active
            self.data_path = file_path
            self.planilha_field.setText(file_path)

            # Atualiza o combo de operadoras
            self.operadora_combo.clear()
            operadoras = set()
            for row in self.sheet.iter_rows(min_row=2, values_only=True):
                if row[3]:  # Coluna 3: OPERADORA
                    operadoras.add(row[3])
            self.operadora_combo.addItems(operadoras)

            # Atualiza o caminho da planilha na instância de MainApp
            self.parent.data_path = file_path
            self.parent.save_settings()  # Salva as configurações
        except Exception as e:
            self.log_message(f"Erro: {str(e)}", area="tecnico")
            QMessageBox.critical(self, "Erro", f"Erro ao carregar a planilha: {e}")

    def select_save_directory(self):
        last_save_dir = QSettings(self.parent.settings_path, QSettings.Format.IniFormat).value("last_save_dir", "")
        dir_path = QFileDialog.getExistingDirectory(self, "Selecionar Diretório de Salvamento", last_save_dir)

        if dir_path and os.path.isdir(dir_path):
            self.save_directory = dir_path
            self.save_dir_field.setText(dir_path)
            
            # Atualiza o diretório de salvamento na instância de MainApp
            self.parent.save_directory = dir_path
            self.parent.save_settings()  # Salva as configurações
        else:
            self.save_directory = ""

    def toggle_automation(self):
        """Alterna entre iniciar e parar a automação."""
        if self.confirm_button.text() == "Iniciar automação":
            self.start_automation()
            self.confirm_button.setText("Parar automação")
            self.confirm_button.clicked.disconnect()  # Remove o conector anterior
            self.confirm_button.clicked.connect(self.toggle_automation)  # Reconecta ao mesmo método
        else:
            self.stop_automation()
            self.confirm_button.setText("Iniciar automação")
            self.confirm_button.clicked.disconnect()  # Remove o conector anterior
            self.confirm_button.clicked.connect(self.toggle_automation)  # Reconecta ao mesmo método
        
    def start_automation(self):
        """Inicia o processo de automação."""
        selected_operadora = self.operadora_combo.currentText()

        if not self.save_directory or not os.path.isdir(self.save_directory):
            QMessageBox.warning(self, "Erro", "Selecione um diretório de salvamento primeiro.")
            return

        if not selected_operadora:
            QMessageBox.warning(self, "Erro", "Selecione uma operadora.")
            return

        if not self.data_path or not os.path.isfile(self.data_path):
            QMessageBox.warning(self, "Erro", "Selecione uma planilha de dados primeiro.")
            return

        if selected_operadora.upper() == "BLUME":
            try:
                self.automator = Blume(self, self.data_path)  # Salva o automator como atributo da classe
            except Exception as e:
                self.log_message(f"Erro: {str(e)}", area="tecnico")
                QMessageBox.critical(self, "Erro", f"Erro ao criar o automator Blume: {e}")
                return
        else:
            QMessageBox.warning(self, "Erro", f"Automação para a operadora {selected_operadora} não está implementada.")
            return

        try:
            # Cria a tarefa de automação
            task = AutomationTask(self.automator, self.get_user_data(selected_operadora), self.log_message)
            
            # Inicia a tarefa no thread pool
            self.threadpool.start(task)
        except Exception as e:
            self.log_message(f"Erro: {str(e)}", area="tecnico")
            QMessageBox.critical(self, "Erro", f"Erro ao iniciar a tarefa de automação: {e}")

    def stop_automation(self):
        """Para a automação em execução."""
        if hasattr(self, 'automator'):
            stop_automation = StopAutomation(self.automator)
            stop_automation.stop()
            self.confirm_button.setText("Iniciar automação")  # Garante que o botão volte ao estado inicial
        else:
            QMessageBox.warning(self, "Erro", "Nenhuma automação em execução para interromper.")

    def get_user_data(self, selected_operadora):
        """Filtra os dados da operadora selecionada."""
        user_data = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[3] == selected_operadora and row[11] != 'COLETADO IA':  # Coluna 3: OPERADORA, Coluna 11: STATUS
                user_data.append({
                    "FORNECEDOR": row[0],
                    "REFERÊNCIA": row[1],
                    "CLIENTE": row[2],
                    "OPERADORA": row[3],
                    "IDENTIFICAÇÃO": row[4],
                    "CÓDIGO": row[5],
                    "PA": row[6],
                    "INDENTIFICAÇÃO INTERNA": row[7],
                    "LOGIN": row[8],
                    "SENHA": row[9],
                    "VENCIMENTO": row[10],
                    "STATUS": row[11],
                    "NOMENCLATURA": row[12]
                })
        return user_data

    def log_message(self, message, area="tecnico", color=None):
        """Exibe mensagens nos logs adequados, com suporte para cores."""
        if area == "tecnico":
            log_area = self.log_tecnico_area
        elif area == "faturas":
            log_area = self.faturas_coletadas_area
        else:
            return  # Área de log inválida

        # Aplica a cor ao texto, se especificada
        if color:
            message = f'<span style="color: {color};">{message}</span>'

        # Adiciona a mensagem ao log
        log_area.append(message)

        # Move o cursor para o final do log
        cursor = log_area.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)
        log_area.setTextCursor(cursor)