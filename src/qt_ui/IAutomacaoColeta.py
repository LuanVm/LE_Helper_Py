from pathlib import Path
from openpyxl import load_workbook
from PyQt6.QtCore import QThreadPool, QSettings
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QGridLayout, QLabel, QLineEdit, QPushButton,
    QComboBox, QTextEdit, QFileDialog
)
from PyQt6.QtGui import QTextCursor

from utils.sheetStyles import (
    estilo_combo_box_light, estilo_hover,
    estilo_log_light, estilo_label_light, campo_qline_light, 
    estilo_log_dark, campo_qline_dark, estilo_combo_box_dark,
    estilo_label_dark
)
from services.AutomacaoColeta import TarefaAutomacao, PararAutomacao

class PainelAutomacaoColeta(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.planilha = None
        self.caminho_dados = ""
        self.pasta_salvamento = ""
        self.threads = QThreadPool()
        self.current_task = None
        self.inicializar_interface()
        self.carregar_configuracoes()

    def apply_styles(self, dark_mode):
        estilo_combo = estilo_combo_box_dark() if dark_mode else estilo_combo_box_light()
        estilo_campo = campo_qline_dark() if dark_mode else campo_qline_light()
        estilo_log = estilo_log_dark() if dark_mode else estilo_log_light()
        estilo_label = estilo_label_dark() if dark_mode else estilo_label_light()

        for btn in (self.botao_pasta, self.botao_planilha, self.botao_iniciar):
            estilo_hover(btn, dark_mode)

        self.combo_operadora.setStyleSheet(estilo_combo)
        self.campo_pasta.setStyleSheet(estilo_campo)
        self.campo_planilha.setStyleSheet(estilo_campo)
        self.log_tecnico.setStyleSheet(estilo_log)
        self.log_faturas.setStyleSheet(estilo_log)
        self.rotulo_pasta.setStyleSheet(estilo_label)
        self.rotulo_planilha.setStyleSheet(estilo_label)
        self.rotulo_operadora.setStyleSheet(estilo_label)

    def inicializar_interface(self):
        self.setWindowTitle("Automação Blume")
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        grid = QGridLayout()
        self.rotulo_pasta = QLabel("Pasta de Salvamento:")
        self.campo_pasta = QLineEdit(); self.campo_pasta.setReadOnly(True)
        self.botao_pasta = QPushButton("Selecionar Pasta")
        self.botao_pasta.clicked.connect(self.selecionar_pasta)

        self.rotulo_planilha = QLabel("Planilha de Dados:")
        self.campo_planilha = QLineEdit(); self.campo_planilha.setReadOnly(True)
        self.botao_planilha = QPushButton("Selecionar Arquivo")
        self.botao_planilha.clicked.connect(self.selecionar_planilha)

        self.rotulo_operadora = QLabel("Operadora:")
        self.combo_operadora = QComboBox()
        self.combo_operadora.addItem("Selecione uma planilha primeiro")

        self.botao_iniciar = QPushButton("Iniciar Automação")
        self.botao_iniciar.clicked.connect(self.alternar_automacao)

        grid.addWidget(self.rotulo_pasta, 0, 0)
        grid.addWidget(self.campo_pasta, 0, 1)
        grid.addWidget(self.botao_pasta, 0, 2)
        grid.addWidget(self.rotulo_planilha, 1, 0)
        grid.addWidget(self.campo_planilha, 1, 1)
        grid.addWidget(self.botao_planilha, 1, 2)
        grid.addWidget(self.rotulo_operadora, 2, 0)
        grid.addWidget(self.combo_operadora, 2, 1)
        grid.addWidget(self.botao_iniciar, 2, 2)

        self.log_tecnico = QTextEdit(); self.log_tecnico.setReadOnly(True)
        self.log_tecnico.setPlaceholderText("Logs técnicos...")
        self.log_faturas = QTextEdit(); self.log_faturas.setReadOnly(True)
        self.log_faturas.setPlaceholderText("Faturas coletadas...")

        layout.addLayout(grid)
        layout.addWidget(self.log_tecnico)
        layout.addWidget(self.log_faturas)

    def carregar_configuracoes(self):
        cfg = QSettings("config.ini", QSettings.Format.IniFormat)
        self.pasta_salvamento = cfg.value("pasta_salvamento", "")
        self.caminho_dados = cfg.value("caminho_dados", "")
        self.campo_pasta.setText(self.pasta_salvamento)
        self.campo_planilha.setText(self.caminho_dados)

    def salvar_configuracoes(self):
        cfg = QSettings("config.ini", QSettings.Format.IniFormat)
        cfg.setValue("pasta_salvamento", self.pasta_salvamento)
        cfg.setValue("caminho_dados", self.caminho_dados)

    def selecionar_pasta(self):
        pasta = QFileDialog.getExistingDirectory(self, "Selecionar Pasta de Salvamento")
        if pasta:
            self.pasta_salvamento = pasta
            self.campo_pasta.setText(pasta)
            self.salvar_configuracoes()
            self.log_mensagem(f"📁 Pasta: {pasta}", "tecnico", "#4CAF50")

    def selecionar_planilha(self):
        caminho, _ = QFileDialog.getOpenFileName(self, "Selecionar Planilha", "", "Excel (*.xlsx *.xlsm)")
        if caminho:
            self.carregar_planilha(caminho)

    def carregar_planilha(self, caminho):
        try:
            wb = load_workbook(caminho)
            sheet = wb.active
            headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
            esperado = [
                "FORNECEDOR","REFERÊNCIA","CLIENTE","OPERADORA",
                "IDENTIFICAÇÃO","CÓDIGO","PA","INDENTIFICAÇÃO INTERNA",
                "LOGIN","SENHA","VENCIMENTO","STATUS","NOMENCLATURA"
            ]
            if headers[:len(esperado)] != esperado:
                raise ValueError("Layout inválido. Esperado: " + ", ".join(esperado))
            operadoras = sorted({row[3] for row in sheet.iter_rows(min_row=2, values_only=True) if row[3]})
            self.combo_operadora.clear()
            self.combo_operadora.addItems(operadoras)
            self.caminho_dados = caminho
            self.campo_planilha.setText(caminho)
            self.salvar_configuracoes()
            self.log_mensagem(f"✅ Operadoras carregadas: {len(operadoras)}", "tecnico", "#4CAF50")
        except Exception as e:
            self.log_mensagem(f"❌ Erro ao carregar planilha: {e}", "tecnico", "#f44336")

    def alternar_automacao(self):
        if self.botao_iniciar.text().startswith("Iniciar"):
            if self.iniciar_automacao():
                self.botao_iniciar.setText("Parar Automação")
        else:
            self.parar_automacao()
            self.botao_iniciar.setText("Iniciar Automação")

    def iniciar_automacao(self) -> bool:
        if not self.validar_campos():
            return False

        dados = self.obter_dados_usuario(self.combo_operadora.currentText())
        try:
            from services.AutomacaoColeta import Blume
            automator = Blume(self, self.caminho_dados)
            self.current_task = TarefaAutomacao(automator, dados, self.log_mensagem)
            self.threads.start(self.current_task)
            self.log_mensagem("🚀 Automação iniciada!", "tecnico", "#FF9800")
            return True
        except Exception as e:
            self.log_mensagem(f"⛔ Falha ao iniciar automação: {e}", "tecnico", "#f44336")
            return False

    def parar_automacao(self):
        if self.current_task:
            PararAutomacao(self.current_task.automator).parar()
            self.log_mensagem("⏹️ Automação interrompida", "tecnico", "#9E9E9E")

    def validar_campos(self) -> bool:
        erros = []
        if not self.pasta_salvamento:
            erros.append("Selecione uma pasta de salvamento")
        if not self.caminho_dados:
            erros.append("Selecione uma planilha de dados")
        if self.combo_operadora.count() == 0:
            erros.append("Selecione uma operadora válida")

        if erros:
            self.log_mensagem("⚠️ Erros encontrados:", "tecnico", "#FFC107")
            for erro in erros:
                self.log_mensagem(f"• {erro}", "tecnico", "#FFC107")
            return False
        return True

    def obter_dados_usuario(self, operadora: str) -> list[dict]:
        wb = load_workbook(self.caminho_dados, data_only=True)
        sheet = wb.active
        dados = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[3] == operadora and row[11] != "COLETADO IA":
                dados.append({
                    "FORNECEDOR": row[0], "REFERÊNCIA": row[1], "CLIENTE": row[2],
                    "OPERADORA": row[3], "IDENTIFICAÇÃO": row[4], "CÓDIGO": row[5],
                    "PA": row[6], "INDENTIFICAÇÃO INTERNA": row[7],
                    "LOGIN": row[8], "SENHA": row[9],
                    "VENCIMENTO": row[10], "STATUS": row[11],
                    "NOMENCLATURA": row[12]
                })
        self.log_mensagem(f"📊 {len(dados)} registros para processamento", "tecnico", "#2196F3")
        return dados

    def log_mensagem(self, mensagem: str, area="tecnico", cor=None):
        target = self.log_tecnico if area == "tecnico" else self.log_faturas
        html = f'<span style="color:{cor}">{mensagem}</span>' if cor else mensagem
        target.append(html)
        cursor = target.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)
        target.setTextCursor(cursor)
        target.ensureCursorVisible()
