import os
from openpyxl import load_workbook
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QGridLayout, QPushButton, QLineEdit,
    QFileDialog, QProgressBar, QTableWidget, QTableWidgetItem,
    QLabel, QHBoxLayout, QHeaderView, QTextEdit,
    QAbstractItemView, QDialog, QCheckBox, QScrollArea
)
from PyQt6.QtCore import QThread, pyqtSlot, QSettings
from openpyxl.utils import get_column_letter

from utils.sheetStyles import (
    estilo_label_light, estilo_label_dark,
    campo_qline_light, campo_qline_dark,
    estilo_tabela_dark, estilo_tabela_light,
    estilo_progress_bar_light, estilo_progress_bar_dark,
    estilo_log_light, estilo_log_dark,
    estilo_hover
)
from services.MesclaPlanilhas import PlanilhaMesclagemWorker


class PainelMesclaPlanilha(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.is_dark_mode = False
        self.colunas_base = set()
        self.cancelar_processo = False
        self.worker_thread = None
        self.init_ui()
        
    def init_ui(self):
        self.setLayout(QVBoxLayout())
        self.layout().setContentsMargins(15, 15, 15, 15)
        self.layout().setSpacing(15)

        self._create_file_controls()
        self._create_table()
        self._create_progress_bar()
        self._create_log_area()
        self._create_action_buttons()

    def _create_file_controls(self):
        grid = QGridLayout()
        grid.setVerticalSpacing(10)
        grid.setHorizontalSpacing(15)
        grid.setColumnStretch(1, 1)

        self.label_pasta = QLabel("Pasta com Arquivos:")
        self.text_pasta = QLineEdit()
        self.text_pasta.setReadOnly(True)
        self.btn_selecionar_pasta = QPushButton("Selecionar Pasta")
        self.btn_selecionar_pasta.clicked.connect(self.selecionar_pasta)

        self.label_base = QLabel("Arquivo Base:")
        self.text_arquivo_base = QLineEdit()
        self.text_arquivo_base.setReadOnly(True)
        self.btn_selecionar_base = QPushButton("Selecionar Base")
        self.btn_selecionar_base.clicked.connect(self.selecionar_arquivo_base)

        self.label_saida = QLabel("Nome do Arquivo Mesclado:")
        self.text_nome_saida = QLineEdit()
        self.text_nome_saida.setPlaceholderText("planilha_mesclada")

        grid.addWidget(self.label_pasta, 0, 0)
        grid.addWidget(self.text_pasta, 0, 1)
        grid.addWidget(self.btn_selecionar_pasta, 0, 2)
        
        grid.addWidget(self.label_base, 1, 0)
        grid.addWidget(self.text_arquivo_base, 1, 1)
        grid.addWidget(self.btn_selecionar_base, 1, 2)
        
        grid.addWidget(self.label_saida, 2, 0)
        grid.addWidget(self.text_nome_saida, 2, 1, 1, 1)

        self.layout().addLayout(grid)

    def _create_table(self):
        self.tabela_arquivos = QTableWidget(0, 2)
        self.tabela_arquivos.setHorizontalHeaderLabels(["Arquivo", "Status"])
        self.tabela_arquivos.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.tabela_arquivos.verticalHeader().hide()
        self.tabela_arquivos.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.layout().addWidget(self.tabela_arquivos)

    def _create_progress_bar(self):
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.layout().addWidget(self.progress_bar)

    def _create_log_area(self):
        self.text_log = QTextEdit()
        self.text_log.setReadOnly(True)
        self.text_log.setPlaceholderText("Log de processamento...")
        self.layout().addWidget(self.text_log)

    def _create_action_buttons(self):
        container = QHBoxLayout()
        container.addStretch()
        
        self.btn_mesclar = QPushButton("Iniciar Mesclagem")
        self.btn_mesclar.setFixedSize(160, 32)
        
        self.btn_cancelar = QPushButton("Cancelar")
        self.btn_cancelar.setFixedSize(160, 32)
        self.btn_cancelar.setEnabled(False)
        
        self.btn_mesclar.clicked.connect(self.iniciar_mesclagem)
        self.btn_cancelar.clicked.connect(self.cancelar_mesclagem)
        
        container.addWidget(self.btn_mesclar)
        container.addWidget(self.btn_cancelar)
        
        self.layout().addLayout(container)

    def apply_styles(self, is_dark_mode):
        self.is_dark_mode = is_dark_mode
        
        label_style = estilo_label_dark() if is_dark_mode else estilo_label_light()
        line_style = campo_qline_dark() if is_dark_mode else campo_qline_light()
        progress_style = estilo_progress_bar_dark() if is_dark_mode else estilo_progress_bar_light()
        table_style = estilo_tabela_dark() if is_dark_mode else estilo_tabela_light()
        log_style = estilo_log_dark() if is_dark_mode else estilo_log_light()

        for label in [self.label_pasta, self.label_base, self.label_saida]:
            label.setStyleSheet(label_style)
            
        for line_edit in [self.text_pasta, self.text_arquivo_base, self.text_nome_saida]:
            line_edit.setStyleSheet(line_style)
            
        self.progress_bar.setStyleSheet(progress_style)
        self.tabela_arquivos.setStyleSheet(table_style)
        self.text_log.setStyleSheet(log_style)

        for button in [self.btn_selecionar_pasta, self.btn_selecionar_base, 
                    self.btn_mesclar, self.btn_cancelar]:
            estilo_hover(button, is_dark_mode)

    def append_log(self, mensagem):
        color = "#e0e0e0" if self.is_dark_mode else "#333333"
        self.text_log.append(f'<span style="color: {color}">{mensagem}</span>')
        self._scroll_to_bottom()

    def _scroll_to_bottom(self):
        self.text_log.verticalScrollBar().setValue(
            self.text_log.verticalScrollBar().maximum()
        )

    def selecionar_pasta(self):
        settings = QSettings("LivreEscolha", "LE_Helper")
        pasta = QFileDialog.getExistingDirectory(self, "Selecionar Pasta", settings.value("last_merge_dir", ""))
        if pasta:
            self.text_pasta.setText(pasta)
            settings.setValue("last_merge_dir", pasta)
            self.append_log(f"📂 Pasta selecionada: {os.path.basename(pasta)}")
            self.carregar_arquivos_pasta(pasta)

    def selecionar_arquivo_base(self):
        settings = QSettings("LivreEscolha", "LE_Helper")
        arquivo, _ = QFileDialog.getOpenFileName(
            self, "Selecionar Arquivo Base", 
            settings.value("last_base_dir", ""), 
            "Excel Files (*.xlsx *.xls)"
        )
        if arquivo:
            self.text_arquivo_base.setText(arquivo)
            settings.setValue("last_base_dir", os.path.dirname(arquivo))
            self.append_log(f"📄 Arquivo base selecionado: {os.path.basename(arquivo)}")
            self.ler_colunas_base(arquivo)

    def carregar_arquivos_pasta(self, pasta):
        self.tabela_arquivos.setRowCount(0)
        arquivos = [f for f in os.listdir(pasta) if f.endswith(('.xlsx', '.xls'))]
        for arquivo in arquivos:
            row = self.tabela_arquivos.rowCount()
            self.tabela_arquivos.insertRow(row)
            self.tabela_arquivos.setItem(row, 0, QTableWidgetItem(arquivo))
            self.tabela_arquivos.setItem(row, 1, QTableWidgetItem("Pendente"))
        self.append_log(f"📑 {len(arquivos)} arquivos encontrados na pasta")

    def ler_colunas_base(self, caminho):
        try:
            wb = load_workbook(caminho, read_only=True)
            sheet = wb.active
            cabecalhos = [cell.value for cell in sheet[1]]
            
            dialogo = DialogoSelecaoColunas(cabecalhos, self)
            if dialogo.exec() == QDialog.DialogCode.Accepted:
                self.colunas_base = set(dialogo.colunas_selecionadas())
                self.append_log(f"🔖 Colunas base selecionadas: {len(self.colunas_base)} colunas")
                
        except Exception as e:
            self.append_log(f"⛔ Erro ao ler arquivo base: {str(e)}")
        finally:
            wb.close()

    @pyqtSlot()
    def iniciar_mesclagem(self):
        if not self.validar_campos():
            return
            
        arquivos = [
            os.path.join(self.text_pasta.text(), self.tabela_arquivos.item(row, 0).text())
            for row in range(self.tabela_arquivos.rowCount())
        ]
        
        self.worker_thread = QThread()
        self.worker = PlanilhaMesclagemWorker(
            arquivos,
            self.text_pasta.text(),
            self.text_nome_saida.text() or "planilha_mesclada",
            self.colunas_base
        )
        
        self.worker.moveToThread(self.worker_thread)
        self.worker_thread.started.connect(self.worker.executar_mesclagem)
        
        self.worker.progress.connect(self.progress_bar.setValue)
        self.worker.atualizar_status.connect(self.atualizar_status_arquivo)
        self.worker.concluido.connect(self.processamento_concluido)
        self.worker.erro.connect(self.mostrar_erro)
        
        self.btn_mesclar.setEnabled(False)
        self.btn_cancelar.setEnabled(True)
        self.worker_thread.start()
        self.append_log("⏳ Iniciando processo de mesclagem...")

    def validar_campos(self):
        if not self.text_pasta.text():
            self.append_log("⚠️ Selecione uma pasta contendo os arquivos!")
            return False
        if not self.colunas_base:
            self.append_log("⚠️ Selecione as colunas base!")
            return False
        return True

    @pyqtSlot(int, str)
    def atualizar_status_arquivo(self, linha, status):
        self.tabela_arquivos.item(linha, 1).setText(status)
        self.append_log(f"📝 Processando {self.tabela_arquivos.item(linha, 0).text()}: {status}")

    @pyqtSlot(str)
    def processamento_concluido(self, mensagem):
        self.worker_thread.quit()
        self.worker_thread.wait()
        self.btn_mesclar.setEnabled(True)
        self.btn_cancelar.setEnabled(False)
        self.append_log(f"✅ {mensagem}")

    @pyqtSlot(str)
    def mostrar_erro(self, mensagem):
        self.append_log(f"⛔ {mensagem}")
        self.processamento_concluido("Processamento interrompido!")

    @pyqtSlot()
    def cancelar_mesclagem(self):
        if self.worker_thread and self.worker_thread.isRunning():
            self.worker.cancelar()
            self.btn_cancelar.setEnabled(False)
            self.worker_thread.quit()
            self.worker_thread.wait()
            self.append_log("🛑 Mesclagem cancelada pelo usuário")


class DialogoSelecaoColunas(QDialog):
    def __init__(self, cabecalhos, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Selecionar Colunas")
        self.setMinimumSize(400, 300)
        
        layout = QVBoxLayout()
        self.scroll = QScrollArea()
        self.widget = QWidget()
        self.layout_colunas = QVBoxLayout(self.widget)
        
        self.checkboxes = []
        for idx, cabecalho in enumerate(cabecalhos):
            cb = QCheckBox(f"{cabecalho} (Coluna {get_column_letter(idx + 1)})")
            cb.setChecked(True)
            self.checkboxes.append((cb, idx))
            self.layout_colunas.addWidget(cb)
        
        self.scroll.setWidget(self.widget)
        self.scroll.setWidgetResizable(True)
        
        btn_confirmar = QPushButton("Confirmar")
        btn_confirmar.clicked.connect(self.accept)
        
        layout.addWidget(self.scroll)
        layout.addWidget(btn_confirmar)
        self.setLayout(layout)

    def colunas_selecionadas(self):
        return [idx for cb, idx in self.checkboxes if cb.isChecked()]