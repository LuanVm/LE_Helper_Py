from PyQt6.QtWidgets import (
    QLabel, QComboBox, QWidget, QGridLayout, QMessageBox,
    QTextEdit, QFileDialog, QLineEdit, QVBoxLayout, QPushButton 
)
from PyQt6.QtCore import QThreadPool, QSettings
from PyQt6.QtGui import QTextCursor
from openpyxl import load_workbook

from utils.GerenEstilos import (
    estilo_combo_box_light, estilo_hover,
    estilo_log_light, estilo_label_light, campo_qline_light, 
    estilo_log_dark, campo_qline_dark, estilo_combo_box_dark,
    estilo_label_dark
)
from services.AutomacaoColeta import TarefaAutomacao, PararAutomacao, Blume

class InterfaceAutoBlume(QWidget):
    """Classe que constrói e gerencia a interface gráfica"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.planilha = None
        self.caminho_dados = ""
        self.pasta_salvamento = ""
        self.threads = QThreadPool()
        self.inicializar_interface()
        self.carregar_configuracoes()

    def apply_styles(self, dark_mode):
        """Aplica estilos dinamicamente com base no tema"""
        estilo_combo = estilo_combo_box_dark() if dark_mode else estilo_combo_box_light()
        estilo_campo = campo_qline_dark() if dark_mode else campo_qline_light()
        estilo_log = estilo_log_dark() if dark_mode else estilo_log_light()
        estilo_label = estilo_label_dark() if dark_mode else estilo_label_light()

        for btn in [self.botao_pasta, self.botao_planilha, self.botao_iniciar]:
            estilo_hover(btn, dark_mode)

        # Aplica estilos nos componentes
        self.combo_operadora.setStyleSheet(estilo_combo)
        self.campo_pasta.setStyleSheet(estilo_campo)
        self.campo_planilha.setStyleSheet(estilo_campo)
        self.log_tecnico.setStyleSheet(estilo_log)
        self.log_faturas.setStyleSheet(estilo_log)
        self.rotulo_pasta.setStyleSheet(estilo_label)
        self.rotulo_planilha.setStyleSheet(estilo_label)
        self.rotulo_operadora.setStyleSheet(estilo_label)

    def inicializar_interface(self):
        """Configura todos os elementos da interface gráfica"""
        self.setWindowTitle("Automação Blume")

        # Layout principal
        layout_principal = QVBoxLayout(self)
        layout_principal.setContentsMargins(10, 10, 10, 10)
        layout_principal.setSpacing(10)

        # Seção de configurações
        layout_superior = QGridLayout()
        
        # Campo de pasta de salvamento
        self.rotulo_pasta = QLabel("Pasta de Salvamento:")
        self.rotulo_pasta.setStyleSheet(estilo_label_light())
        self.campo_pasta = QLineEdit()
        self.campo_pasta.setReadOnly(True)
        self.campo_pasta.setStyleSheet(campo_qline_light())
        self.botao_pasta = QPushButton("Selecionar Pasta")
        estilo_hover(self.botao_pasta)
        self.botao_pasta.clicked.connect(self.selecionar_pasta)

        # Campo de planilha de dados
        self.rotulo_planilha = QLabel("Planilha de Dados:")
        self.rotulo_planilha.setStyleSheet(estilo_label_light())
        self.campo_planilha = QLineEdit()
        self.campo_planilha.setReadOnly(True)
        self.campo_planilha.setStyleSheet(campo_qline_light())
        self.botao_planilha = QPushButton("Selecionar Arquivo")
        estilo_hover(self.botao_planilha)
        self.botao_planilha.clicked.connect(self.selecionar_planilha)

        # Seleção de operadora
        self.rotulo_operadora = QLabel("Operadora:")
        self.rotulo_operadora.setStyleSheet(estilo_label_light())
        self.combo_operadora = QComboBox()
        self.combo_operadora.addItem("Selecione uma planilha primeiro")
        self.combo_operadora.setStyleSheet(estilo_combo_box_light())

        # Botão de controle
        self.botao_iniciar = QPushButton("Iniciar Automação")
        estilo_hover(self.botao_iniciar)
        self.botao_iniciar.clicked.connect(self.alternar_automacao)

        # Adicionando elementos ao layout
        layout_superior.addWidget(self.rotulo_pasta, 0, 0)
        layout_superior.addWidget(self.campo_pasta, 0, 1)
        layout_superior.addWidget(self.botao_pasta, 0, 2)
        
        layout_superior.addWidget(self.rotulo_planilha, 1, 0)
        layout_superior.addWidget(self.campo_planilha, 1, 1)
        layout_superior.addWidget(self.botao_planilha, 1, 2)
        
        layout_superior.addWidget(self.rotulo_operadora, 2, 0)
        layout_superior.addWidget(self.combo_operadora, 2, 1)
        layout_superior.addWidget(self.botao_iniciar, 2, 2)

        # Área de logs
        self.log_tecnico = QTextEdit()
        self.log_tecnico.setPlaceholderText("Logs técnicos...")
        self.log_tecnico.setReadOnly(True)
        self.log_tecnico.setStyleSheet(estilo_log_light())
        
        self.log_faturas = QTextEdit()
        self.log_faturas.setPlaceholderText("Faturas coletadas...")
        self.log_faturas.setReadOnly(True)
        self.log_faturas.setStyleSheet(estilo_log_light())

        # Montagem final
        layout_principal.addLayout(layout_superior)
        layout_principal.addWidget(self.log_tecnico)
        layout_principal.addWidget(self.log_faturas)

    def carregar_configuracoes(self):
        """Carrega as configurações salvas anteriormente"""
        config = QSettings("config.ini", QSettings.Format.IniFormat)
        self.pasta_salvamento = config.value("pasta_salvamento", "")
        self.caminho_dados = config.value("caminho_dados", "")
        self.campo_pasta.setText(self.pasta_salvamento)

    def salvar_configuracoes(self):
        """Salva as configurações atuais"""
        config = QSettings("config.ini", QSettings.Format.IniFormat)
        config.setValue("pasta_salvamento", self.pasta_salvamento)
        config.setValue("caminho_dados", self.caminho_dados)

    def selecionar_pasta(self):
        """Seleciona a pasta para salvar os boletos"""
        pasta = QFileDialog.getExistingDirectory(self, "Selecionar Pasta de Salvamento")
        if pasta:
            self.pasta_salvamento = pasta
            self.campo_pasta.setText(pasta)
            self.salvar_configuracoes()

    def selecionar_planilha(self):
        """Seleciona o arquivo Excel com os dados"""
        arquivo, _ = QFileDialog.getOpenFileName(
            self,
            "Selecionar Planilha",
            "",
            "Arquivos Excel (*.xlsx *.xlsm)"
        )
        if arquivo:
            self.carregar_planilha(arquivo)

    def carregar_planilha(self, caminho):
        """Carrega e processa o arquivo Excel selecionado"""
        try:
            self.planilha = load_workbook(caminho).active
            self.caminho_dados = caminho
            self.campo_planilha.setText(caminho)
            self.salvar_configuracoes()

            operadoras = set()
            for linha in self.planilha.iter_rows(min_row=2, values_only=True):
                if linha[3]:
                    operadoras.add(linha[3])
            
            self.combo_operadora.clear()
            self.combo_operadora.addItems(operadoras)
            
        except Exception as erro:
            self.log_mensagem(f"Erro ao carregar planilha: {erro}", area="tecnico")
            QMessageBox.critical(self, "Erro", f"Falha ao carregar arquivo: {erro}")

    def alternar_automacao(self):
        """Controla o início/parada da automação"""
        if self.botao_iniciar.text() == "Iniciar Automação":
            self.iniciar_automacao()
            self.botao_iniciar.setText("Parar Automação")
        else:
            self.parar_automacao()
            self.botao_iniciar.setText("Iniciar Automação")

    def iniciar_automacao(self):
        """Inicia o processo de automação"""
        if not self.validar_campos():
            return

        operadora = self.combo_operadora.currentText()
        dados = self.obter_dados_usuario(operadora)

        try:
            self.automator = Blume(self, self.caminho_dados)
            tarefa = TarefaAutomacao(self.automator, dados, self.log_mensagem)
            self.threads.start(tarefa)
        except Exception as erro:
            self.log_mensagem(f"Falha ao iniciar: {erro}", area="tecnico")

    def parar_automacao(self):
        """Interrompe a automação em execução"""
        if hasattr(self, 'automator'):
            PararAutomacao(self.automator).parar()
        self.botao_iniciar.setText("Iniciar Automação")

    def validar_campos(self):
        """Valida os campos antes de iniciar"""
        erros = []
        if not self.pasta_salvamento:
            erros.append("Selecione uma pasta de salvamento")
        if not self.caminho_dados:
            erros.append("Selecione uma planilha de dados")
        if self.combo_operadora.currentText() == "Selecione uma planilha primeiro":
            erros.append("Selecione uma operadora válida")
        
        if erros:
            QMessageBox.warning(self, "Atenção", "\n".join(erros))
            return False
        return True

    def obter_dados_usuario(self, operadora):
        """Extrai os dados da planilha para a operadora selecionada"""
        dados = []
        for linha in self.planilha.iter_rows(min_row=2, values_only=True):
            if linha[3] == operadora and linha[11] != 'COLETADO IA':
                dados.append({
                    "FORNECEDOR": linha[0],
                    "REFERÊNCIA": linha[1],
                    "CLIENTE": linha[2],
                    "OPERADORA": linha[3],
                    "IDENTIFICAÇÃO": linha[4],
                    "CÓDIGO": linha[5],
                    "PA": linha[6],
                    "INDENTIFICAÇÃO INTERNA": linha[7],
                    "LOGIN": linha[8],
                    "SENHA": linha[9],
                    "VENCIMENTO": linha[10],
                    "STATUS": linha[11],
                    "NOMENCLATURA": linha[12]
                })
        return dados

    def log_mensagem(self, mensagem, area="tecnico", cor=None):
        """Exibe mensagens nas áreas de log correspondentes"""
        area_log = self.log_tecnico if area == "tecnico" else self.log_faturas

        if cor:
            mensagem = f'<span style="color: {cor};">{mensagem}</span>'
        area_log.append(mensagem)
        cursor = area_log.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)
        area_log.setTextCursor(cursor)