import os
import logging
import time
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QGridLayout, QPushButton, QLineEdit,
    QFileDialog, QProgressBar, QTableWidget, QTableWidgetItem,
    QMessageBox, QLabel, QHBoxLayout, QHeaderView,
    QAbstractItemView, QDialog, QCheckBox, QScrollArea, QVBoxLayout
)
from PyQt6.QtCore import QThread, pyqtSignal, pyqtSlot, QSettings, QObject
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from GerenEstilos import (
    estilo_label_light, estilo_label_dark,
    campo_qline_light, campo_qline_dark,
    estilo_tabela_dark, estilo_tabela_light,
    estilo_progress_bar_light, estilo_progress_bar_dark,
    estilo_hover
)

#### TASKS ####

#### O app está em um loop infinito após finalizar, não se sabe se concluiu a mesclagem
#### melhorar design da tabela

class PlanilhaMesclagemWorker(QObject):
    progress = pyqtSignal(int)
    concluido = pyqtSignal(str)  # Agora envia o caminho do arquivo
    erro = pyqtSignal(str)
    atualizar_status = pyqtSignal(int, str)

    def __init__(self, arquivos, pasta_saida, nome_arquivo, colunas_selecionadas):
        super().__init__()
        self.arquivos = arquivos
        self.pasta_saida = pasta_saida
        self.nome_arquivo = nome_arquivo
        self.colunas_selecionadas = sorted(colunas_selecionadas)
        self._cancelar = False
        self.estilos_base = None
        self.larguras_colunas = None
        self.wb_saida = None
        self.ws_saida = None
        self.caminho_saida = None

    def _carregar_estilos_base(self, arquivo_base):
        """Carrega os estilos da planilha base de forma segura"""
        try:
            wb = load_workbook(arquivo_base)
            ws = wb.active
            
            self.larguras_colunas = {}
            self.estilos_base = {}
            
            for col in self.colunas_selecionadas:
                col_letter = get_column_letter(col + 1)
                
                # Largura da coluna
                if col_letter in ws.column_dimensions:
                    self.larguras_colunas[col_letter] = ws.column_dimensions[col_letter].width
                
                # Estilo do cabeçalho
                cell = ws.cell(row=1, column=col+1)
                if cell.style:
                    self.estilos_base[col] = cell.style
            
            wb.close()
        except Exception as e:
            logging.error(f"Erro ao carregar estilos base: {str(e)}")

    def _aplicar_estilos(self):
        """Aplica os estilos carregados na planilha de saída de forma segura"""
        try:
            if self.larguras_colunas:
                for col, width in self.larguras_colunas.items():
                    if width and col in self.ws_saida.column_dimensions:
                        self.ws_saida.column_dimensions[col].width = width

            if self.estilos_base and self.ws_saida.max_row >= 1:
                for col_idx, style in self.estilos_base.items():
                    if col_idx < len(self.ws_saida[1]):
                        self.ws_saida.cell(row=1, column=col_idx+1).style = style
        except Exception as e:
            logging.error(f"Erro ao aplicar estilos: {str(e)}")

    def _salvar_parcialmente(self):
        """Salva o arquivo de forma segura durante o cancelamento"""
        try:
            if self.wb_saida:
                # Cria nome único para o arquivo parcial
                partial_name = f"{self.nome_arquivo}_PARCIAL_{int(time.time())}.xlsx"
                self.caminho_saida = os.path.join(self.pasta_saida, partial_name)
                self.wb_saida.save(self.caminho_saida)
                return self.caminho_saida
        except Exception as e:
            logging.error(f"Erro ao salvar parcialmente: {str(e)}")
        return None

    def executar_mesclagem(self):
        try:
            self.caminho_saida = os.path.join(self.pasta_saida, f"{self.nome_arquivo}.xlsx")
            self.wb_saida = Workbook()
            self.ws_saida = self.wb_saida.active

            # Carrega estilos do primeiro arquivo
            if self.arquivos:
                self._carregar_estilos_base(self.arquivos[0])
                self._aplicar_estilos()

            total_arquivos = len(self.arquivos)
            for idx, arquivo in enumerate(self.arquivos):
                if self._cancelar:
                    break

                self.atualizar_status.emit(idx, "Processando...")
                
                try:
                    wb_entrada = load_workbook(arquivo, read_only=True, data_only=True)
                    ws_entrada = wb_entrada.active
                    
                    for row in ws_entrada.iter_rows(min_row=2, values_only=True):
                        if self._cancelar:
                            break
                            
                        nova_linha = []
                        for col in self.colunas_selecionadas:
                            try:
                                valor = row[col] if col < len(row) else ""
                                valor = "" if isinstance(valor, str) and valor.startswith("#REF!") else valor
                                nova_linha.append(valor)
                            except IndexError:
                                nova_linha.append("")
                        
                        self.ws_saida.append(nova_linha)
                    
                    wb_entrada.close()
                    self.progress.emit(int((idx + 1) / total_arquivos * 100))
                    self.atualizar_status.emit(idx, "Concluído")

                except Exception as e:
                    self.atualizar_status.emit(idx, f"Erro: {str(e)[:30]}")
                    logging.error(f"Erro no arquivo {arquivo}: {str(e)}")

            if self._cancelar:
                caminho_parcial = self._salvar_parcialmente()
                self.concluido.emit(f"Processo cancelado. Arquivo parcial salvo em: {caminho_parcial}" if caminho_parcial 
                                  else "Cancelado mas houve erro ao salvar")
            else:
                self._aplicar_estilos()
                self.wb_saida.save(self.caminho_saida)
                self.concluido.emit(f"Arquivo final salvo em: {self.caminho_saida}")

        except Exception as e:
            self.erro.emit(f"Erro crítico: {str(e)}")
            logging.error("Erro na mesclagem", exc_info=True)
        finally:
            if self.wb_saida:
                self.wb_saida.close()

    def cancelar(self):
        self._cancelar = True

class DialogoSelecaoColunas(QDialog):
    def __init__(self, cabecalhos, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Selecionar Colunas")
        self.setMinimumSize(400, 300)
        
        layout = QVBoxLayout()
        self.scroll = QScrollArea()
        self.widget = QWidget()
        self.layout_colunas = QVBoxLayout(self.widget)
        
        self.checkboxes = []
        for idx, cabecalho in enumerate(cabecalhos):
            cb = QCheckBox(f"{cabecalho} (Coluna {get_column_letter(idx + 1)})")
            cb.setChecked(True)
            self.checkboxes.append((cb, idx))
            self.layout_colunas.addWidget(cb)
        
        self.scroll.setWidget(self.widget)
        self.scroll.setWidgetResizable(True)
        
        btn_confirmar = QPushButton("Confirmar")
        btn_confirmar.clicked.connect(self.accept)
        
        layout.addWidget(self.scroll)
        layout.addWidget(btn_confirmar)
        self.setLayout(layout)

    def colunas_selecionadas(self):
        return [idx for cb, idx in self.checkboxes if cb.isChecked()]

class PainelMesclaPlanilha(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.is_dark_mode = False
        self.colunas_base = set()
        self.cancelar_processo = False
        self.worker_thread = None
        self.init_ui()
        
    def init_ui(self):
        self.setLayout(QVBoxLayout())
        self.layout().setContentsMargins(15, 15, 15, 15)
        self.layout().setSpacing(15)

        self._create_file_controls()
        self._create_table()
        self._create_progress_bar()
        self._create_action_buttons()

    def _create_file_controls(self):
        grid = QGridLayout()
        grid.setVerticalSpacing(10)
        grid.setHorizontalSpacing(15)
        grid.setColumnStretch(1, 1)

        # Pasta de entrada
        self.label_pasta = QLabel("Pasta com Arquivos:")
        self.text_pasta = QLineEdit()
        self.text_pasta.setReadOnly(True)
        self.btn_selecionar_pasta = QPushButton("Selecionar Pasta")
        self.btn_selecionar_pasta.clicked.connect(self.selecionar_pasta)

        # Arquivo base
        self.label_base = QLabel("Arquivo Base:")
        self.text_arquivo_base = QLineEdit()
        self.text_arquivo_base.setReadOnly(True)
        self.btn_selecionar_base = QPushButton("Selecionar Base")
        self.btn_selecionar_base.clicked.connect(self.selecionar_arquivo_base)

        # Nome do arquivo de saída
        self.label_saida = QLabel("Nome do Arquivo Mesclado:")
        self.text_nome_saida = QLineEdit()
        self.text_nome_saida.setPlaceholderText("planilha_mesclada")

        # Posicionamento dos elementos
        grid.addWidget(self.label_pasta, 0, 0)
        grid.addWidget(self.text_pasta, 0, 1)
        grid.addWidget(self.btn_selecionar_pasta, 0, 2)
        
        grid.addWidget(self.label_base, 1, 0)
        grid.addWidget(self.text_arquivo_base, 1, 1)
        grid.addWidget(self.btn_selecionar_base, 1, 2)
        
        grid.addWidget(self.label_saida, 2, 0)
        grid.addWidget(self.text_nome_saida, 2, 1, 1, 2)

        self.layout().addLayout(grid)

    def _create_table(self):
        self.tabela_arquivos = QTableWidget(0, 2)
        self.tabela_arquivos.setHorizontalHeaderLabels(["Arquivo", "Status"])
        self.tabela_arquivos.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.tabela_arquivos.verticalHeader().hide()
        self.tabela_arquivos.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.layout().addWidget(self.tabela_arquivos)

    def _create_progress_bar(self):
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.layout().addWidget(self.progress_bar)

    def _create_action_buttons(self):
        container = QHBoxLayout()
        container.setContentsMargins(0, 0, 0, 0)
        container.setSpacing(10)
        
        # Espaço expansível à esquerda
        container.addStretch()
        
        # Botões com tamanho fixo e alinhamento à direita
        self.btn_mesclar = QPushButton("Iniciar Mesclagem")
        self.btn_mesclar.setFixedSize(160, 32)
        
        self.btn_cancelar = QPushButton("Cancelar")
        self.btn_cancelar.setFixedSize(160, 32)
        self.btn_cancelar.setEnabled(False)
        
        # Conexões de sinais
        self.btn_mesclar.clicked.connect(self.iniciar_mesclagem)
        self.btn_cancelar.clicked.connect(self.cancelar_mesclagem)
        
        container.addWidget(self.btn_mesclar)
        container.addWidget(self.btn_cancelar)
        
        self.layout().addLayout(container)

    def apply_styles(self, is_dark_mode):
        self.is_dark_mode = is_dark_mode
        
        # Aplica estilos individuais
        label_style = estilo_label_dark() if is_dark_mode else estilo_label_light()
        line_style = campo_qline_dark() if is_dark_mode else campo_qline_light()
        progress_style = estilo_progress_bar_dark() if is_dark_mode else estilo_progress_bar_light()
        table_style = estilo_tabela_dark() if is_dark_mode else estilo_tabela_light()

        # Aplica nos componentes
        for label in [self.label_pasta, self.label_base, self.label_saida]:
            label.setStyleSheet(label_style)
            
        for line_edit in [self.text_pasta, self.text_arquivo_base, self.text_nome_saida]:
            line_edit.setStyleSheet(line_style)
            
        self.progress_bar.setStyleSheet(progress_style)
        self.tabela_arquivos.setStyleSheet(table_style)

        # Configurações adicionais para a tabela
        if is_dark_mode:
            self.tabela_arquivos.setAlternatingRowColors(True)
            self.tabela_arquivos.setStyleSheet(table_style + """
                QTableWidget {
                    alternate-background-color: #353535;
                }
            """)
            header_style = """
                QHeaderView::section {
                    background-color: #3d3d3d;
                    color: #e0e0e0;
                }
            """
        else:
            self.tabela_arquivos.setAlternatingRowColors(True)
            self.tabela_arquivos.setStyleSheet(table_style + """
                QTableWidget {
                    alternate-background-color: #f8f8f8;
                }
            """)
            header_style = """
                QHeaderView::section {
                    background-color: #f4f4f4;
                    color: #1c1c1c;
                }
            """
        
        self.tabela_arquivos.horizontalHeader().setStyleSheet(header_style)

        # Estilo dos botões
        for button in [self.btn_selecionar_pasta, self.btn_selecionar_base, 
                    self.btn_mesclar, self.btn_cancelar]:
            estilo_hover(button, is_dark_mode)

    def selecionar_pasta(self):
        settings = QSettings("LivreEscolha", "LE_Helper")
        last_dir = settings.value("last_merge_dir", "")
        
        pasta = QFileDialog.getExistingDirectory(self, "Selecionar Pasta", last_dir)
        if pasta:
            self.text_pasta.setText(pasta)
            settings.setValue("last_merge_dir", pasta)
            self.carregar_arquivos_pasta(pasta)

    def selecionar_arquivo_base(self):
        settings = QSettings("LivreEscolha", "LE_Helper")
        last_dir = settings.value("last_base_dir", "")
        
        arquivo, _ = QFileDialog.getOpenFileName(
            self, "Selecionar Arquivo Base", last_dir, 
            "Excel Files (*.xlsx *.xls)"
        )
        if arquivo:
            self.text_arquivo_base.setText(arquivo)
            settings.setValue("last_base_dir", os.path.dirname(arquivo))
            self.ler_colunas_base(arquivo)

    def carregar_arquivos_pasta(self, pasta):
        self.tabela_arquivos.setRowCount(0)
        arquivos = [f for f in os.listdir(pasta) if f.endswith(('.xlsx', '.xls'))]
        for arquivo in arquivos:
            row = self.tabela_arquivos.rowCount()
            self.tabela_arquivos.insertRow(row)
            self.tabela_arquivos.setItem(row, 0, QTableWidgetItem(arquivo))
            self.tabela_arquivos.setItem(row, 1, QTableWidgetItem("Pendente"))

    def ler_colunas_base(self, caminho):
        try:
            wb = load_workbook(caminho, read_only=True)
            sheet = wb.active
            cabecalhos = [cell.value for cell in sheet[1]]
            
            dialogo = DialogoSelecaoColunas(cabecalhos, self)
            if dialogo.exec() == QDialog.DialogCode.Accepted:
                self.colunas_base = set(dialogo.colunas_selecionadas())
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao ler arquivo base:\n{str(e)}")
        finally:
            wb.close()

    @pyqtSlot()
    def iniciar_mesclagem(self):
        if not self.validar_campos():
            return
            
        arquivos = [
            os.path.join(self.text_pasta.text(), self.tabela_arquivos.item(row, 0).text())
            for row in range(self.tabela_arquivos.rowCount())
        ]
        
        self.worker_thread = QThread()
        self.worker = PlanilhaMesclagemWorker(
            arquivos,
            self.text_pasta.text(),
            self.text_nome_saida.text() or "planilha_mesclada",
            self.colunas_base
        )
        
        self.worker.moveToThread(self.worker_thread)
        self.worker_thread.started.connect(self.worker.executar_mesclagem)
        
        self.worker.progress.connect(self.progress_bar.setValue)
        self.worker.atualizar_status.connect(self.atualizar_status_arquivo)
        self.worker.concluido.connect(self.processamento_concluido)
        self.worker.erro.connect(self.mostrar_erro)
        
        self.btn_mesclar.setEnabled(False)
        self.btn_cancelar.setEnabled(True)
        self.worker_thread.start()

    def validar_campos(self):
        if not self.text_pasta.text():
            QMessageBox.warning(self, "Aviso", "Selecione uma pasta contendo os arquivos!")
            return False
        if not self.colunas_base:
            QMessageBox.warning(self, "Aviso", "Selecione as colunas base!")
            return False
        return True

    @pyqtSlot(int, str)
    def atualizar_status_arquivo(self, linha, status):
        self.tabela_arquivos.item(linha, 1).setText(status)

    @pyqtSlot(str)
    def processamento_concluido(self, mensagem):
        self.worker_thread.quit()
        self.worker_thread.wait()
        self.btn_mesclar.setEnabled(True)
        self.btn_cancelar.setEnabled(False)
        self.btn_cancelar.setText("Cancelar")
        QMessageBox.information(self, "Status", mensagem)

    @pyqtSlot(str)
    def mostrar_erro(self, mensagem):
        QMessageBox.critical(self, "Erro", mensagem)
        self.processamento_concluido()

    @pyqtSlot()
    def cancelar_mesclagem(self):
        if self.worker_thread and self.worker_thread.isRunning():
            self.worker.cancelar()
            self.btn_cancelar.setEnabled(False)
            self.btn_cancelar.setText("Cancelando...")
            self.worker_thread.quit()
            self.worker_thread.wait()

class DialogoSelecaoColunas(QDialog):
    def __init__(self, cabecalhos, parent=None):
        super().__init__(parent)
        # Filtra cabeçalhos None e ajusta os índices
        self.cabecalhos_validos = [(i, str(h) if h is not None else f"Coluna {i+1}") 
                                 for i, h in enumerate(cabecalhos) if h is not None]
        
        self.setWindowTitle("Selecionar Colunas")
        self.setMinimumSize(400, 300)
        
        layout = QVBoxLayout()
        self.scroll = QScrollArea()
        self.widget = QWidget()
        self.layout_colunas = QVBoxLayout(self.widget)
        
        self.checkboxes = []
        for idx, (original_idx, cabecalho) in enumerate(self.cabecalhos_validos):
            cb = QCheckBox(f"{cabecalho} (Coluna {get_column_letter(original_idx + 1)})")
            cb.setChecked(True)
            self.checkboxes.append((cb, original_idx))  # Mantém o índice original
            self.layout_colunas.addWidget(cb)
        
        self.scroll.setWidget(self.widget)
        self.scroll.setWidgetResizable(True)
        
        btn_confirmar = QPushButton("Confirmar")
        btn_confirmar.clicked.connect(self.accept)
        
        layout.addWidget(self.scroll)
        layout.addWidget(btn_confirmar)
        self.setLayout(layout)

    def colunas_selecionadas(self):
        return [original_idx for cb, original_idx in self.checkboxes if cb.isChecked()]