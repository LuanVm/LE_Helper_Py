import re
from openpyxl.utils import datetime as xl_datetime
from datetime import time as dt_time, datetime as dt_datetime
import gc
import logging
import os
import unicodedata
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QGridLayout, QPushButton, QLineEdit,
    QFileDialog, QProgressBar, QTextEdit, QCheckBox,
    QMessageBox, QLabel, QHBoxLayout, QApplication
)
from PyQt6.QtCore import QThread, pyqtSignal, pyqtSlot, QSettings, Qt
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, NamedStyle

# Importa os estilos visuais
from GerenEstilos import (
    estilo_label_light, estilo_label_dark,
    campo_qline_light, campo_qline_dark,
    estilo_check_box_light, estilo_check_box_dark,
    estilo_log_light, estilo_log_dark,
    estilo_progress_bar_light, estilo_progress_bar_dark,
    estilo_hover
)

class PainelProcessamentoAgitel(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.is_dark_mode = False
        self.init_ui()

    def init_ui(self):
        self.setLayout(QVBoxLayout())
        self.layout().setContentsMargins(15, 15, 15, 15)
        self.layout().setSpacing(10)

        self._create_file_controls()
        self._create_progress_bar()
        self._create_results_area()

    def _create_file_controls(self):
        grid = QGridLayout()
        grid.setVerticalSpacing(10)
        grid.setHorizontalSpacing(15)
        grid.setColumnStretch(1, 1)

        self.label_file = QLabel("Arquivo XLSX (Planilha da Agitel):")
        self.text_file = QLineEdit()
        self.text_file.setReadOnly(True)

        self.btn_select_file = QPushButton("Selecionar Arquivo")
        self.btn_select_file.setFixedSize(160, 32)
        self.btn_process = QPushButton("Processar")
        self.btn_process.setFixedSize(160, 32)
        self.checkbox_equalize = QCheckBox("Equalizar 'Região'")

        button_layout = QHBoxLayout()
        button_layout.addWidget(self.btn_select_file)
        button_layout.addWidget(self.btn_process)
        button_layout.setSpacing(10)

        grid.addWidget(self.label_file, 0, 0)
        grid.addWidget(self.text_file, 0, 1)
        grid.addLayout(button_layout, 0, 2)
        grid.addWidget(self.checkbox_equalize, 0, 3, Qt.AlignmentFlag.AlignLeft)

        self.btn_select_file.clicked.connect(self.select_file)
        self.btn_process.clicked.connect(self.process_file)

        self.layout().addLayout(grid)

    def _create_progress_bar(self):
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.layout().addWidget(self.progress_bar)

    def _create_results_area(self):
        self.text_results = QTextEdit()
        self.text_results.setReadOnly(True)
        self.text_results.setPlaceholderText("Resultados do processamento...")
        self.layout().addWidget(self.text_results)

    def apply_styles(self, is_dark_mode):
        self.is_dark_mode = is_dark_mode
        styles = {
            'label': estilo_label_dark() if is_dark_mode else estilo_label_light(),
            'line': campo_qline_dark() if is_dark_mode else campo_qline_light(),
            'check': estilo_check_box_dark() if is_dark_mode else estilo_check_box_light(),
            'log': estilo_log_dark() if is_dark_mode else estilo_log_light(),
            'progress': estilo_progress_bar_dark() if is_dark_mode else estilo_progress_bar_light()
        }

        self.label_file.setStyleSheet(styles['label'])
        self.text_file.setStyleSheet(styles['line'])
        self.checkbox_equalize.setStyleSheet(styles['check'])
        self.text_results.setStyleSheet(styles['log'])
        self.progress_bar.setStyleSheet(styles['progress'])

        for btn in [self.btn_select_file, self.btn_process]:
            estilo_hover(btn, is_dark_mode)

    def select_file(self):
        settings = QSettings("LivreEscolha", "LE_Helper")
        last_dir = settings.value("last_open_dir", "")
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Selecionar Arquivo Excel", last_dir, "Excel Files (*.xlsx)"
        )
        if file_path:
            self.text_file.setText(file_path)
            settings.setValue("last_open_dir", os.path.dirname(file_path))

    def process_file(self):
        file_path = self.text_file.text()
        if not self._validate_file(file_path):
            return

        self.progress_bar.reset()
        self.btn_process.setEnabled(False)
        
        self.processor = ProcessarArquivo(file_path, self.checkbox_equalize.isChecked())
        self.processor.progress.connect(self.update_progress)
        self.processor.finished.connect(self.on_process_finished)
        self.processor.error.connect(self.show_error)
        self.processor.update_log.connect(self.append_log)
        self.processor.start()

    def _validate_file(self, file_path):
        if not file_path:
            QMessageBox.warning(self, "Aviso", "Selecione um arquivo Excel.")
            return False
        if not os.path.isfile(file_path):
            QMessageBox.warning(self, "Aviso", "Arquivo inválido.")
            return False
        return True

    @pyqtSlot(int)
    def update_progress(self, value):
        self.progress_bar.setValue(value)

    @pyqtSlot(str)
    def on_process_finished(self, message):
        self.btn_process.setEnabled(True)
        self.append_log(message)

    @pyqtSlot(str)
    def append_log(self, message):
        color = "#e0e0e0" if self.is_dark_mode else "#333333"
        self.text_results.append(f'<span style="color: {color}">{message}</span>')
        QApplication.processEvents()

    @pyqtSlot(str)
    def show_error(self, message):
        QMessageBox.critical(self, "Erro", message)
        self.append_log("Processamento interrompido devido a erro crítico")

class ProcessarArquivo(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(str)
    error = pyqtSignal(str)
    update_log = pyqtSignal(str)

    COLUNAS_SAIDA = [
        'Data', 'Origem', 'Serviço', 'Região', 
        'Destino', 'Duração', 'Duração (minutos)', 'Valor'
    ]

    def __init__(self, file_path, equalize):
        super().__init__()
        self.file_path = file_path
        self.equalize = equalize
        self._setup_styles()

    def _setup_styles(self):
        self.styles = {
            'date': NamedStyle(name="date", number_format='YYYY-MM-DD HH:MM:SS'),
            'minutes': NamedStyle(name="minutes", number_format='0.0'),
            'currency': NamedStyle(name="currency", number_format='R$ #,##0.00'),
            'duration': NamedStyle(name="duration", number_format='hh:mm:ss')  # Novo estilo
        }

    def run(self):
        try:
            wb = load_workbook(self.file_path, read_only=True)
            valid_sheets = []
            batch_counter = 0

            # Passo 1: Identificar se a primeira aba é Resumo
            primeira_aba = wb.worksheets[0].title.lower() if wb.worksheets else ""
            ignorar_primeira = "resumo" in primeira_aba

            # Passo 2: Processar abas (ignorando apenas a primeira se for Resumo)
            for sheet in wb.worksheets:
                # Pula primeira aba se for Resumo
                if ignorar_primeira and sheet == wb.worksheets[0]:
                    continue
                    
                header_row = self._find_header_row(sheet)
                if header_row:
                    valid_sheets.append(sheet)
                else:
                    self.update_log.emit(f"Aviso: {sheet.title} ignorada (cabeçalho não encontrado)")

            total_sheets = len(valid_sheets)
            progress_per_sheet = 100 / total_sheets if total_sheets > 0 else 0

            output_wb = Workbook()
            output_sheet = output_wb.active
            self._create_header(output_wb, output_sheet)

            # Passo 3: Processar todas as abas válidas
            for index, sheet in enumerate(valid_sheets, 1):
                if self.isInterruptionRequested():
                    break

                self.update_log.emit(f"Processando: {sheet.title}")
                
                # Processa todas as linhas
                for chunk in self._process_sheet(sheet):
                    for row in chunk:
                        output_sheet.append(row)

                # Atualização do progresso
                current_progress = int(index * progress_per_sheet)
                self.progress.emit(min(current_progress, 100))

                # Limpeza periódica
                batch_counter += 1
                if batch_counter >= 10 or index == total_sheets:
                    self._clean_and_sort_output(output_sheet)
                    batch_counter = 0

            # Passo 4: Finalizações
            self._apply_final_formatting(output_sheet)
            if self.equalize:
                self._equalize_regiao(output_sheet)

            output_path = self._get_output_path()
            output_wb.save(output_path)
            self.finished.emit(f"Arquivo salvo em: {output_path}")
            self.progress.emit(100)

        except Exception as e:
            self.error.emit(f"Erro crítico: {str(e)}")
            logging.exception("Erro durante o processamento")
        finally:
            if 'wb' in locals(): wb.close()
            if 'output_wb' in locals(): output_wb.close()
            gc.collect()

    def _clean_and_sort_output(self, sheet):
        """Remove linhas vazias e classifica por Região"""
        try:
            # Coleta e filtra dados
            all_rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(cell not in (None, "", 0) for cell in row):
                    all_rows.append(row)

            # Classificação por Região (coluna 3)
            all_rows.sort(key=lambda x: str(x[3]).lower() if x[3] else "")

            # Limpa a planilha e reinsere os dados
            sheet.delete_rows(2, sheet.max_row)
            for row in all_rows:
                sheet.append(row)

        except Exception as e:
            self.update_log.emit(f"Erro na organização: {str(e)[:50]}")

    def _apply_final_formatting(self, sheet):
        for row in sheet.iter_rows(min_row=2):
            cell = row[5]  # Coluna F (Duração)
            cell.number_format = 'hh:mm:ss'

    def _process_sheet(self, sheet):
        header_row = self._find_header_row(sheet)
        if not header_row:
            return

        indices = self._get_column_indices(header_row)
        start_row = header_row[0].row + 1
        chunk = []

        rows = list(sheet.iter_rows(min_row=start_row, values_only=True))
        total_rows = len(rows)
        current_index = 0

        while current_index < total_rows:
            if self.isInterruptionRequested():
                return

            row = rows[current_index]
            current_index += 1

            processed_row = self._process_row(row, indices)
            if processed_row:
                chunk.append(processed_row)
        if chunk:
            yield chunk

    def _process_row(self, row, indices):
        try:
            # Conversões seguras com fallback
            data = self._convert_date(row[indices.get('data', -1)]) or ""
            origem = str(row[indices.get('origem', -1)] or "").strip()
            destino = str(row[indices.get('destino', -1)] or "").strip()
            
            return [
                data,
                origem,
                str(row[indices.get('servico', -1)] or ""),
                str(row[indices.get('regiao', -1)] or ""),
                destino,
                self._convert_duration(row[indices.get('duracao', -1)]),
                self._duration_to_minutes(row[indices.get('duracao', -1)]),
                self._parse_currency(row[indices.get('preco', -1)])
            ]
        except Exception as e:
            self.update_log.emit(f"Linha ignorada: {str(e)[:50]}")
            return None

    def _create_header(self, wb, sheet):
        sheet.append(self.COLUNAS_SAIDA)
        for col, title in enumerate(self.COLUNAS_SAIDA, 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Formatação inicial da coluna Duração
        sheet.column_dimensions['F'].number_format = 'hh:mm:ss'

    def _get_output_path(self):
        base, ext = os.path.splitext(self.file_path)
        return f"{base}_leitura_agitel{ext}"

    def _convert_date(self, value):
        if isinstance(value, dt_datetime):
            return xl_datetime.to_excel(value)
        return value

    def _convert_duration(self, value):
        if isinstance(value, dt_time):
            return value.hour/24 + value.minute/1440 + value.second/86400
        elif isinstance(value, str):
            try:
                h, m, s = map(int, value.split(':'))
                return h/24 + m/1440 + s/86400
            except:
                return 0.0
        return value

    def _duration_to_minutes(self, value):
        if isinstance(value, dt_time):
            return round(value.hour * 60 + value.minute + value.second / 60, 1)
        return 0.0

    def _parse_currency(self, value):
        try:
            return float(str(value).replace('R$', '').replace(',', '.').strip())
        except:
            return 0.0

    def _equalize_regiao(self, sheet):
        for row in sheet.iter_rows(min_row=2):
            cell = row[3]
            if not cell.value:
                continue
                
            valor = str(cell.value).lower()
            if "fixo" in valor:
                cell.value = "Fixo"
            elif any(x in valor for x in ["movel", "móvel"]):  # Aceita ambas as formas
                cell.value = "Móvel"  # Padroniza com acento
            elif not valor.strip():
                cell.value = "Intragrupo"

    def _find_header_row(self, sheet):
        """Encontra a linha de cabeçalho com verificação de padrão de dados"""
        essential_columns = {'data', 'origem', 'destino', 'duracao', 'preco'}
        
        for row in sheet.iter_rows(max_row=20):  # Verifica até a linha 20
            # Ignora linhas que parecem conter dados
            if self._is_data_row(row):
                continue
                
            # Verifica colunas essenciais
            found = set()
            for cell in row:
                if cell.value:
                    normalized = self._normalize(str(cell.value))
                    if normalized in essential_columns:
                        found.add(normalized)
            
            if found >= essential_columns:
                return row
        return None
    
    def _is_data_row(self, row):
        """Verifica se a linha contém padrão de dados (ex: datas/números)"""
        data_patterns = 0
        for cell in row:
            if isinstance(cell.value, (dt_datetime, int, float)):
                data_patterns += 1
            elif isinstance(cell.value, str):
                if re.match(r"\d{2}/\d{2}/\d{4}", cell.value):  # Data
                    data_patterns += 1
                elif re.match(r"\d{2}:\d{2}:\d{2}", cell.value):  # Duração
                    data_patterns += 1
                elif "R$" in cell.value:  # Valor monetário
                    data_patterns += 1
        
        return data_patterns >= 3  # Se encontrar 3+ padrões, é linha de dados

    def _get_column_indices(self, header_row):
        headers = [self._normalize(str(cell.value)) for cell in header_row]
        mapping = {
            'data': ['data', 'datachamada', 'datahora', 'datahorario'],
            'origem': ['origem', 'ramalorigem', 'setor', 'operador'],
            'servico': ['servico', 'tiposervico', 'serviço', 'tipochamada'],
            'regiao': ['regiao', 'região', 'localchamada', 'ddd'],
            'destino': ['destino', 'numerodestino', 'telefone', 'ramaldestino'],
            'duracao': ['duracao', 'tempochamada', 'tempogasto', 'duraçao'],
            'preco': ['preco', 'custochamada', 'valor', 'tarifa']
        }

        indices = {}
        for key, aliases in mapping.items():
            for alias in aliases:
                normalized = self._normalize(alias)
                if normalized in headers:
                    indices[key] = headers.index(normalized)
                    break
            else:
                raise ValueError(f"Coluna '{key}' não encontrada")
        return indices

    def _normalize(self, text):
        return ''.join(c for c in unicodedata.normalize('NFD', str(text).lower()) 
                      if not unicodedata.combining(c))