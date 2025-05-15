import os
import time
import shutil
import re
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook
from PyQt6.QtCore import QRunnable, QMutex
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from PyPDF2 import PdfReader

class TarefaAutomacao(QRunnable):
    def __init__(self, automator, dados_usuario, funcao_log):
        super().__init__()
        self.automator = automator
        self.dados_usuario = dados_usuario
        self.funcao_log = funcao_log

    def run(self):
        try:
            self.funcao_log("Iniciando automação...", area="tecnico")
            if hasattr(self.automator, 'flag_parar') and self.automator.flag_parar:
                self.funcao_log("Automação cancelada antes de iniciar", area="tecnico")
                return
            self.automator.executar_automacao(self.dados_usuario)
            self.funcao_log("Automação concluída!", area="tecnico")
        except Exception as erro:
            self.funcao_log(f"Erro durante a automação: {str(erro)}", area="tecnico")

class PararAutomacao:
    def __init__(self, automator):
        self.automator = automator

    def parar(self):
        if hasattr(self.automator, 'flag_parar'):
            self.automator.flag_parar = True
            self.automator.parent.log_mensagem("Parando automação...", area="tecnico", cor="red")
            self.fechar_navegadores()
        else:
            self.automator.parent.log_mensagem("Nenhuma automação em execução", area="tecnico")

    def fechar_navegadores(self):
        if hasattr(self.automator, 'drivers') and self.automator.drivers:
            self.automator.parent.log_mensagem("Fechando navegadores...", area="tecnico")
            for driver in self.automator.drivers:
                try:
                    driver.quit()
                except Exception as erro:
                    self.automator.parent.log_mensagem(f"Erro ao fechar navegador: {erro}", area="tecnico")
            self.automator.drivers.clear()

class Blume:
    """Classe principal que implementa a automação para a operadora Blume"""
    
    def __init__(self, parent, caminho_dados):
        self.parent = parent
        self.caminho_dados = caminho_dados
        self.diretorio_download = os.path.join(os.path.expanduser('~'), 'Downloads')  # ✅ aqui
        self.planilha = load_workbook(caminho_dados, data_only=True).active
        self.mutex = QMutex()
        self.flag_parar = False
        self.drivers = []

    def inicializar_navegador(self):
        try:
            opcoes = webdriver.ChromeOptions()
            opcoes.add_argument("--disable-extensions")
            opcoes.add_argument("--disable-popup-blocking")
            opcoes.add_argument("--window-size=600,1000")
            #opcoes.add_argument("--headless")
            self.parent.log_mensagem("Abrindo navegador...", area="tecnico")
            driver = webdriver.Chrome(
                service=Service(ChromeDriverManager().install()),
                options=opcoes
            )
            self.drivers.append(driver)
            return driver
        except Exception as erro:
            self.parent.log_mensagem(f"Falha ao iniciar navegador: {erro}", area="tecnico")
            raise

    def executar_automacao(self, dados_usuario):
        self.parent.log_mensagem("Iniciando coleta para Blume...", area="tecnico")
        if self.verificar_coleta_completa():
            self.parent.log_mensagem("Todas faturas já foram coletadas!", area="tecnico")
            return

        logins_indisponiveis_loop_atual = set()

        for usuario in dados_usuario:
            if self.flag_parar:
                self.parent.log_mensagem("Processo interrompido!", area="tecnico")
                self.fechar_navegadores()
                return

            if usuario['STATUS'] == 'COLETADO IA':
                continue

            if usuario['STATUS'] == 'INDISPONIVEL' and usuario['LOGIN'] in logins_indisponiveis_loop_atual:
                continue

            driver = None
            try:
                driver = self.inicializar_navegador()
                wait = WebDriverWait(driver, 2)
                driver.get("https://portal.blumetelecom.com.br")
                self.fazer_login(driver, wait, usuario)
                self.processar_boletos(driver, wait, usuario)
                self.marcar_pendentes_indisponiveis(usuario['LOGIN'])
            except Exception as erro:
                self.parent.log_mensagem(f"Erro no processamento: {str(erro)}", area="tecnico")
                logins_indisponiveis_loop_atual.add(usuario['LOGIN'])
            finally:
                if driver:
                    driver.quit()
                    if driver in self.drivers:
                        self.drivers.remove(driver)

    def fazer_login(self, driver, wait, dados_usuario):
        try:
            self.parent.log_mensagem(f"Realizando login no acesso: {dados_usuario['LOGIN']}", area="tecnico")
            time.sleep(1)
            wait.until(lambda driver: driver.execute_script("return document.readyState") == "complete")

            time.sleep(1)
            campo_email = wait.until(EC.presence_of_element_located((By.ID, "loginUsername")))
            driver.execute_script("arguments[0].scrollIntoView(true);", campo_email)
            campo_email.send_keys(dados_usuario['LOGIN'])

            time.sleep(1)
            campo_senha = wait.until(EC.element_to_be_clickable((By.NAME, "password")))
            driver.execute_script("arguments[0].scrollIntoView(true);", campo_senha)
            campo_senha.clear()
            campo_senha.send_keys(dados_usuario['SENHA'])

            time.sleep(1)
            botao_login = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "MuiButton-label")))
            driver.execute_script("arguments[0].scrollIntoView(true);", botao_login)
            botao_login.click()

            time.sleep(1)
            WebDriverWait(driver, 15).until(EC.url_contains("portal"))
            time.sleep(2)

            driver.get("https://portal.blumetelecom.com.br/billings")
            self.parent.log_mensagem("Login realizado com sucesso!", area="tecnico")
        except Exception as erro:
            self.parent.log_mensagem(f"Falha no login: {erro}", area="tecnico")
            raise

    def processar_boletos(self, driver, wait, dados_usuario):
        try:
            driver.get("https://portal.blumetelecom.com.br/billings")
            time.sleep(1)

            try:
                wait.until(EC.presence_of_element_located(
                    (By.XPATH, "//h5[contains(text(), 'Você não possui faturas em aberto')]")
                ))
                self.parent.log_mensagem("Nenhuma fatura disponível", area="tecnico")
                self.atualizar_status_planilha(dados_usuario['LOGIN'], 'INDISPONIVEL')
                return
            except:
                pass  # Continua o processamento

            botoes_pagamento = wait.until(EC.presence_of_all_elements_located(
                (By.XPATH, "//span[text()='Escolher forma de pagamento']")
            ))
            total_botoes = len(botoes_pagamento)
            self.parent.log_mensagem(f"Foram encontrados {total_botoes} boleto(s) para processar", area="tecnico")

            for indice in range(total_botoes):
                try:
                    botoes_pagamento = wait.until(EC.presence_of_all_elements_located(
                        (By.XPATH, "//span[text()='Escolher forma de pagamento']")
                    ))
                    botao = botoes_pagamento[indice]

                    self.parent.log_mensagem(f"Processando boleto {indice + 1}", area="tecnico")
                    driver.execute_script("arguments[0].click();", botao)
                    time.sleep(1)

                    botao_pagamento_boleto = wait.until(EC.element_to_be_clickable(
                        (By.XPATH, "//p[text()='Pagamento Boleto']")))
                    driver.execute_script("arguments[0].click();", botao_pagamento_boleto)
                    time.sleep(1)

                    botao_baixar_boleto = wait.until(EC.element_to_be_clickable(
                        (By.XPATH, "//p[text()='Baixar boleto']")))
                    driver.execute_script("arguments[0].click();", botao_baixar_boleto)

                    self.parent.log_mensagem("Aguardando finalização do download (10 segundos)...", area="tecnico")
                    time.sleep(10)

                    tempo_limite = datetime.now() - timedelta(seconds=15)
                    arquivos_pdf = list(Path(self.diretorio_download).glob("*.pdf"))
                    arquivos_pdf = [f for f in arquivos_pdf if datetime.fromtimestamp(f.stat().st_mtime) > tempo_limite]

                    if not arquivos_pdf:
                        self.parent.log_mensagem("Nenhum PDF recente encontrado após o download.", area="tecnico")
                        continue

                    ultimo_pdf = max(arquivos_pdf, key=lambda f: f.stat().st_mtime)
                    caminho_pdf = str(ultimo_pdf)
                    contrato = self.extrair_contrato_pdf(caminho_pdf)

                    if self.verificar_pendencia_pdf(caminho_pdf):
                        if contrato:
                            nomenclatura = self.obter_nomenclatura(contrato)
                        else:
                            nomenclatura = dados_usuario['LOGIN']  # Nome com CNPJ

                        nomenclatura = re.sub(r'[<>:"/\\|?*]', '', str(nomenclatura))
                        nome_arquivo = f"{nomenclatura} - PENDENCIA.pdf"
                        destino = os.path.join(self.parent.pasta_salvamento, "Pendencias")
                        os.makedirs(destino, exist_ok=True)
                        destino_final = os.path.join(destino, nome_arquivo)
                        shutil.move(caminho_pdf, destino_final)
                        self.parent.log_mensagem(f"Arquivo movido para pendências: {destino_final}", area="tecnico")
                    elif contrato:
                        self.processar_arquivo_baixado(contrato, caminho_pdf, dados_usuario)
                        self.atualizar_status_planilha(contrato, 'COLETADO IA')
                    else:
                        # Caso não encontrou o contrato e também não é pendência
                        pasta_erro = os.path.join(self.parent.pasta_salvamento, "Boletos não encontrados")
                        os.makedirs(pasta_erro, exist_ok=True)
                        nome_arquivo = f"{dados_usuario['LOGIN']}.pdf"
                        destino = os.path.join(pasta_erro, nome_arquivo)
                        shutil.move(caminho_pdf, destino)
                        self.parent.log_mensagem(f"Arquivo movido para: {destino}", area="tecnico")

                    driver.get("https://portal.blumetelecom.com.br/billings")
                    time.sleep(2)

                except Exception as erro:
                    self.parent.log_mensagem(f"Erro ao processar boleto {indice + 1}: {erro}", area="tecnico")

        except Exception as erro:
            self.parent.log_mensagem(f"Erro geral no processamento: {erro}", area="tecnico")
            raise

    def verificar_pendencia_pdf(self, caminho_pdf):
        """
        Verifica se a fatura é uma pendência com base na data de vencimento explícita.
        Busca especificamente pela string 'Data de Vencimento' seguida de uma data.
        """
        try:
            with open(caminho_pdf, "rb") as f:
                leitor = PdfReader(f)
                texto = "".join(p.extract_text() or "" for p in leitor.pages)

            # Buscar por data explicitamente associada à 'Data de Vencimento'
            padrao = re.compile(r"(\d{2}/\d{2}/\d{4})\s*Data de Vencimento", re.IGNORECASE)
            match = padrao.search(texto)
            if match:
                data_vencimento = datetime.strptime(match.group(1), "%d/%m/%Y").date()
                hoje = datetime.now().date()
                return data_vencimento < hoje

            # Fallback opcional: tentar pegar a primeira data visível (pode ser perigoso)
            self.parent.log_mensagem("Data de vencimento não localizada explicitamente.", area="tecnico")
            return False
        except Exception as erro:
            self.parent.log_mensagem(f"Erro ao verificar pendência no PDF: {erro}", area="tecnico")
            return False

    def processar_boleto_baixado(self, caminho_pdf, dados_usuario):
        """Processa o PDF baixado após confirmação de download concluído"""
        try:
            contrato = self.extrair_contrato_pdf(caminho_pdf)
            if contrato:
                self.processar_arquivo_baixado(contrato, caminho_pdf, dados_usuario)
            else:
                self.mover_arquivo_nao_encontrado(caminho_pdf)
        except Exception as erro:
            self.parent.log_mensagem(f"Erro ao processar PDF: {erro}", area="tecnico")

    def baixar_boleto(self, wait, dados_usuario, indice):
        """Realiza o download e processamento do boleto com verificação baseada em novo arquivo detectado."""
        try:
            time.sleep(1)
            botao_download = wait.until(
                EC.presence_of_element_located((By.XPATH, "//p[text()='Baixar boleto']"))
            )
            botao_download.click()
            time.sleep(5)
            
            self.parent.log_mensagem("Aguardando novo arquivo na pasta de download...", area="tecnico")
            caminho_download = os.path.join(os.path.expanduser('~'), 'Downloads')
            arquivo = self.aguardar_download(caminho_download)

            if arquivo:
                self.parent.log_mensagem(f"Arquivo baixado: {os.path.basename(arquivo)}", area="tecnico")
                contrato = self.extrair_contrato_pdf(arquivo)
                if contrato:
                    self.processar_arquivo_baixado(contrato, arquivo, dados_usuario)
                else:
                    self.mover_arquivo_nao_encontrado(arquivo)
            else:
                self.parent.log_mensagem("Falha ao identificar novo arquivo PDF baixado.", area="tecnico")
        except Exception as erro:
            self.parent.log_mensagem(f"Erro no download: {erro}", area="tecnico")

    def aguardar_download(download_dir, timeout=60, tempo_estabilizacao=2):
        """
        Aguarda a conclusão de downloads na pasta especificada.
        Um download é considerado concluído quando não existem arquivos com a extensão '.crdownload'.
        """
        limite = datetime.now() + timedelta(seconds=timeout)

        while datetime.now() < limite:
            time.sleep(1)
            arquivos = list(Path(download_dir).glob("*.crdownload"))
            if not arquivos:
                # Verifica se os arquivos estão estáveis
                arquivos_pdf = list(Path(download_dir).glob("*.pdf"))
                for arquivo in arquivos_pdf:
                    tamanho_anterior = arquivo.stat().st_size
                    time.sleep(tempo_estabilizacao)
                    tamanho_atual = arquivo.stat().st_size
                    if tamanho_anterior == tamanho_atual:
                        return str(arquivo)
            time.sleep(1)

        return None

    def extrair_contrato_pdf(self, caminho_pdf):
        """Extrai o número do contrato de um arquivo PDF"""
        try:
            time.sleep(1)
            with open(caminho_pdf, "rb") as f:
                leitor = PdfReader(f)
                texto = "".join(pagina.extract_text() or "" for pagina in leitor.pages)

            padrao = re.compile(
                r'Contrato\s*R\$\s*\d{1,4}(?:[.,]\d{3})*[.,]\d{2}\s*0*(\d{1,7})',
                re.IGNORECASE
            )
            match = padrao.search(texto)
            if match:
                contrato = str(int(match.group(1)))
                return contrato

            return None
        except Exception as erro:
            self.parent.log_mensagem(f"Erro na leitura do PDF: {erro}", area="tecnico")
            return None

    def processar_arquivo_baixado(self, contrato, arquivo, dados_usuario):
        """Processa o arquivo baixado e atualiza a planilha"""
        contrato = contrato.lstrip('0')
        # Verifica se o contrato já foi coletado (linha já marcada como 'COLETADO IA')
        for linha in self.planilha.iter_rows(min_row=2, values_only=True):
            if str(linha[4]).lstrip('0') == contrato and linha[11] == 'COLETADO IA':
                self.parent.log_mensagem(f"O contrato {contrato} já havia sido baixado.", area="tecnico")
                os.remove(arquivo)  # Remove o arquivo duplicado
                return
        # Caso contrário, procede com o processamento normal
        if self.verificar_contrato_planilha(contrato):
            nomenclatura = self.obter_nomenclatura(contrato)
            nomenclatura = re.sub(r'[<>:"/\\|?*]', '', str(nomenclatura))
            destino = os.path.join(self.parent.pasta_salvamento, f"{nomenclatura}.pdf")
            self.parent.log_mensagem(f"Movendo arquivo para: {destino}", area="tecnico")
            shutil.move(arquivo, destino)
            self.parent.log_mensagem(nomenclatura, area="faturas")
            self.atualizar_status_planilha(contrato, 'COLETADO IA')
        else:
            self.mover_arquivo_nao_encontrado(arquivo)

    def mover_arquivo_nao_encontrado(self, arquivo):
        """Move arquivos não identificados para a pasta 'Boletos não encontrados'"""
        pasta_erro = os.path.join(self.parent.pasta_salvamento, "Boletos não encontrados")
        os.makedirs(pasta_erro, exist_ok=True)
        destino = os.path.join(pasta_erro, os.path.basename(arquivo))
        shutil.move(arquivo, destino)
        self.parent.log_mensagem(f"Arquivo movido para: {destino}", area="tecnico")

    def verificar_contrato_planilha(self, contrato):
        """Verifica se o contrato existe na planilha"""
        for linha in self.planilha.iter_rows(min_row=2, values_only=True):
            if str(linha[4]).lstrip('0') == contrato:
                return True
        return False

    def obter_nomenclatura(self, contrato):
        """Obtém a nomenclatura correta do contrato"""
        for linha in self.planilha.iter_rows(min_row=2, values_only=True):
            if str(linha[4]).lstrip('0') == contrato:
                return linha[12]
        return None

    def atualizar_status_planilha(self, identificador, status):
        """
        Atualiza o status na planilha Excel para todas as linhas que batem com o identificador.
        O identificador pode ser o LOGIN ou o número do contrato (sem zeros à esquerda).
        """
        self.mutex.lock()
        try:
            for linha in self.planilha.iter_rows(min_row=2):
                if (str(linha[8].value) == identificador or str(linha[4].value).lstrip('0') == identificador):
                    linha[11].value = status
            self.planilha.parent.save(self.caminho_dados)
        finally:
            self.mutex.unlock()

    def marcar_pendentes_indisponiveis(self, login):
        """
        Após o processamento dos boletos para o usuário, percorre as linhas da planilha
        que pertencem ao login informado e marca como "INDISPONIVEL" aquelas que ainda não foram
        atualizadas para "COLETADO IA".
        """
        self.mutex.lock()
        try:
            for linha in self.planilha.iter_rows(min_row=2, values_only=False):
                if str(linha[8].value) == login and linha[11].value not in ['COLETADO IA', 'INDISPONIVEL']:
                    linha[11].value = 'INDISPONIVEL'
            self.planilha.parent.save(self.caminho_dados)
        finally:
            self.mutex.unlock()

    def verificar_coleta_completa(self):
        """
        Verifica se todas as faturas foram coletadas.
        Considera processadas apenas as linhas com status 'COLETADO IA'.
        Faturas com status 'INDISPONIVEL' devem ser reavaliadas futuramente.
        """
        for linha in self.planilha.iter_rows(min_row=2, values_only=True):
            if linha[11] != 'COLETADO IA':
                return False
        return True

    def fechar_navegadores(self):
        """Fecha todas as instâncias do navegador"""
        for driver in self.drivers:
            try:
                driver.quit()
            except:
                pass
        self.drivers.clear()