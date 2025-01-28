import os
import logging
import time
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from PyQt6.QtCore import QObject, pyqtSignal


class PlanilhaMesclagemWorker(QObject):
    progress = pyqtSignal(int)
    concluido = pyqtSignal(str)
    erro = pyqtSignal(str)
    atualizar_status = pyqtSignal(int, str)

    def __init__(self, arquivos, pasta_saida, nome_arquivo, colunas_selecionadas):
        super().__init__()
        self.arquivos = arquivos
        self.pasta_saida = pasta_saida
        self.nome_arquivo = nome_arquivo
        self.colunas_selecionadas = sorted(colunas_selecionadas)
        self._cancelar = False
        self.estilos_base = None
        self.larguras_colunas = None
        self.wb_saida = None
        self.ws_saida = None
        self.caminho_saida = None

    def _carregar_estilos_base(self, arquivo_base):
        """Carrega estilos da planilha base para aplicar na saída"""
        try:
            wb = load_workbook(arquivo_base)
            ws = wb.active
            
            self.larguras_colunas = {}
            self.estilos_base = {}
            
            for col in self.colunas_selecionadas:
                col_letter = get_column_letter(col + 1)
                
                # Obter o valor do cabeçalho da coluna
                nome_coluna = ws.cell(row=1, column=col + 1).value

                # Ignorar colunas com nome "None" ou valor None
                if nome_coluna is None or (str(nome_coluna).strip() == "None"):
                    logging.info(f"Ignorando coluna {col_letter} com nome: {nome_coluna}")
                    continue
                
                # Largura da coluna
                if col_letter in ws.column_dimensions:
                    self.larguras_colunas[col_letter] = ws.column_dimensions[col_letter].width
                
                # Estilo da célula
                cell = ws.cell(row=1, column=col + 1)
                if cell.style:
                    self.estilos_base[col] = cell.style
            
            wb.close()
        except Exception as e:
            logging.error(f"Erro ao carregar estilos base: {str(e)}")

    def _aplicar_estilos(self):
        """Aplica estilos carregados na planilha de saída"""
        try:
            if self.larguras_colunas:
                for col, width in self.larguras_colunas.items():
                    if width and col in self.ws_saida.column_dimensions:
                        self.ws_saida.column_dimensions[col].width = width

            if self.estilos_base and self.ws_saida.max_row >= 1:
                for col_idx, style in self.estilos_base.items():
                    if col_idx < len(self.ws_saida[1]):
                        self.ws_saida.cell(row=1, column=col_idx+1).style = style
        except Exception as e:
            logging.error(f"Erro ao aplicar estilos: {str(e)}")

    def _salvar_parcialmente(self):
        """Salva arquivo parcial em caso de cancelamento"""
        try:
            if self.wb_saida:
                partial_name = f"{self.nome_arquivo}_PARCIAL_{int(time.time())}.xlsx"
                self.caminho_saida = os.path.join(self.pasta_saida, partial_name)
                self.wb_saida.save(self.caminho_saida)
                return self.caminho_saida
        except Exception as e:
            logging.error(f"Erro ao salvar parcialmente: {str(e)}")
        return None

    def executar_mesclagem(self):
        """Método principal que executa o processo de mesclagem"""
        try:
            self.caminho_saida = os.path.join(self.pasta_saida, f"{self.nome_arquivo}.xlsx")
            self.wb_saida = Workbook()
            self.ws_saida = self.wb_saida.active

            if self.arquivos:
                self._carregar_estilos_base(self.arquivos[0])
                self._aplicar_estilos()

            total_arquivos = len(self.arquivos)
            for idx, arquivo in enumerate(self.arquivos):
                if self._cancelar:
                    break

                self.atualizar_status.emit(idx, "Processando...")
                
                try:
                    wb_entrada = load_workbook(arquivo, read_only=True, data_only=True)
                    ws_entrada = wb_entrada.active
                    
                    for row in ws_entrada.iter_rows(min_row=2, values_only=True):
                        if self._cancelar:
                            break
                            
                        nova_linha = [row[col] if col < len(row) else "" for col in self.colunas_selecionadas]
                        self.ws_saida.append(nova_linha)
                    
                    wb_entrada.close()
                    self.progress.emit(int((idx + 1) / total_arquivos * 100))
                    self.atualizar_status.emit(idx, "Concluído")

                except Exception as e:
                    self.atualizar_status.emit(idx, f"Erro: {str(e)[:30]}")
                    logging.error(f"Erro no arquivo {arquivo}: {str(e)}")

            if self._cancelar:
                caminho_parcial = self._salvar_parcialmente()
                self.concluido.emit(f"Processo cancelado. Arquivo parcial salvo em: {caminho_parcial}" if caminho_parcial 
                                  else "Cancelado mas houve erro ao salvar")
            else:
                self._aplicar_estilos()
                self.wb_saida.save(self.caminho_saida)
                self.concluido.emit(f"Arquivo final salvo em: {self.caminho_saida}")

        except Exception as e:
            self.erro.emit(f"Erro crítico: {str(e)}")
            logging.error("Erro na mesclagem", exc_info=True)
        finally:
            if self.wb_saida:
                self.wb_saida.close()

    def cancelar(self):
        """Marca o processo para cancelamento"""
        self._cancelar = True
