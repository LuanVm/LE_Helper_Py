import os
import gc
import logging
from datetime import time as dt_time, datetime as dt_datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import datetime as xl_datetime
from openpyxl.styles import Font, Alignment, NamedStyle
from PyQt6.QtCore import QThread, pyqtSignal
from PyQt6.QtWidgets import QApplication
import unicodedata


class ProcessamentoAgitel(QThread):
    progress_updated = pyqtSignal(int)
    process_finished = pyqtSignal(str)
    error_occurred = pyqtSignal(str)
    log_updated = pyqtSignal(str)

    COLUNAS_SAIDA = [
        'Data', 'Origem', 'Serviço', 'Região',
        'Destino', 'Duração', 'Duração (minutos)', 'Valor'
    ]

    def __init__(self):
        super().__init__()
        self._setup_styles()
        self._file_path = None
        self._equalize = False
        self._log_emitted = set()

    def processar(self, file_path, equalize):
        self._file_path = file_path
        self._equalize = equalize
        self.start()

    def run(self):
        try:
            wb = load_workbook(self._file_path, read_only=True)
            output_wb = Workbook()
            output_sheet = output_wb.active
            self._processar_planilhas(wb, output_wb, output_sheet)
            self._finalizar_processamento(output_wb, output_sheet)
        except Exception as e:
            self._handle_error(e)
        finally:
            self._cleanup_resources(wb, output_wb)

    def _processar_planilhas(self, wb, output_wb, output_sheet):
        MAX_LINHAS = 1048575
        valid_sheets = self._identificar_abas_validas(wb)
        sheet_num = 1
        total_linhas = 0

        # Resetar barra de progresso
        self.progress_updated.emit(0)

        for index, sheet in enumerate(valid_sheets):
            # Criar nova planilha se necessário (com margem de segurança)
            if total_linhas > (MAX_LINHAS - 50000):  # 50k linhas de margem
                sheet_num += 1
                output_sheet = output_wb.create_sheet(title=f"Parte_{sheet_num}")
                self._create_header(output_sheet)
                total_linhas = 0
                self.log_updated.emit(f"Nova planilha criada: Parte_{sheet_num}")

            # Processamento com monitoramento preciso
            linhas_processadas = self._processar_aba(sheet, output_sheet)
            total_linhas += linhas_processadas

            # Atualização de progresso baseada em abas processadas
            progress_per_sheet = 100 / len(valid_sheets)
            self._atualizar_progresso(index + 1, progress_per_sheet)

        # Após processar todas as abas, ordenar ambas as sheets criadas
        self._ordenar_e_mesclar_sheets(output_wb)

    def _ordenar_e_mesclar_sheets(self, output_wb):
        """Ordena e mescla todas as sheets criadas."""
        all_data = []
        for sheet in output_wb.worksheets:
            if sheet.max_row > 1:  # Ignorar sheets sem dados
                header = [cell.value for cell in sheet[1]]
                if not all_data:  # Adicionar cabeçalho apenas uma vez
                    all_data.append(header)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    all_data.append(row)

        # Criar nova sheet para dados mesclados
        merged_sheet = output_wb.create_sheet(title="Dados_Mesclados")
        for row in all_data:
            merged_sheet.append(row)

        # Ordenar pela coluna "Departamento" (índice 3)
        self._clean_and_sort_output(merged_sheet)

    def _identificar_abas_validas(self, wb):
        valid_sheets = []
        primeira_aba = wb.worksheets[0].title.lower() if wb.worksheets else ""
        ignorar_primeira = "resumo" in primeira_aba
        for sheet in wb.worksheets:
            if ignorar_primeira and sheet == wb.worksheets[0]:
                continue
            if self._find_header_row(sheet):
                valid_sheets.append(sheet)
            else:
                self.log_updated.emit(f"Aviso: {sheet.title} ignorada (cabeçalho não encontrado)")
        return valid_sheets

    def _processar_aba(self, sheet, output_sheet):
        try:
            if sheet.title not in self._log_emitted:
                self.log_updated.emit(f"Processando: {sheet.title}")
                self._log_emitted.add(sheet.title)
            linhas_processadas = 0
            header_row = self._find_header_row(sheet)
            if not header_row:
                return 0
            indices = self._get_column_indices(header_row)
            start_row = header_row[0].row + 1
            for row in sheet.iter_rows(min_row=start_row, values_only=True):
                if self.isInterruptionRequested():
                    return linhas_processadas
                processed_row = self._process_row(row, indices)
                if processed_row:
                    output_sheet.append(processed_row)
                    linhas_processadas += 1
            return linhas_processadas
        except Exception as e:
            self.log_updated.emit(f"Erro na aba {sheet.title}: {str(e)[:50]}")
            return linhas_processadas
        
    def _process_row(self, row, indices):
        try:
            data = self._convert_date(row[indices['data']]) or ""
            origem = str(row[indices['origem']] or "").strip()
            destino = str(row[indices['destino']] or "").strip()
            return [
                data,
                origem,
                str(row[indices['servico']] or ""),
                str(row[indices['regiao']] or ""),
                destino,
                self._convert_duration(row[indices['duracao']]),
                self._duration_to_minutes(row[indices['duracao']]),
                self._parse_currency(row[indices['preco']])
            ]
        except Exception as e:
            self.log_updated.emit(f"Linha ignorada: {str(e)[:50]}")
            return None

    def _finalizar_processamento(self, output_wb, output_sheet):
        self._apply_final_formatting(output_sheet)
        if self._equalize:
            self._equalize_regiao(output_sheet)
        output_path = self._get_output_path()
        output_wb.save(output_path)
        self.process_finished.emit(f"Arquivo salvo em: {output_path}")

    def _apply_final_formatting(self, sheet):
        """
        Aplica formatações finais à planilha de saída.
        """
        try:
            # Aplicar estilos às colunas específicas
            for row in sheet.iter_rows(min_row=2):
                # Formatar coluna "Duração" (índice 6)
                duration_cell = row[5]
                duration_cell.number_format = self.styles['duration'].number_format

                # Formatar coluna "Duração (minutos)" (índice 7)
                minutes_cell = row[6]
                minutes_cell.number_format = self.styles['minutes'].number_format

                # Formatar coluna "Valor" (índice 8)
                currency_cell = row[7]
                currency_cell.number_format = self.styles['currency'].number_format

            # Garantir que o cabeçalho tenha estilo consistente
            for col, title in enumerate(self.COLUNAS_SAIDA, 1):
                header_cell = sheet.cell(row=1, column=col)
                header_cell.font = Font(bold=True)
                header_cell.alignment = Alignment(horizontal='center')

            # Ajustar largura das colunas automaticamente
            for col in range(1, len(self.COLUNAS_SAIDA) + 1):
                max_length = 0
                column = sheet.column_dimensions[chr(64 + col)]  # Colunas A, B, C...
                for cell in sheet[chr(64 + col)]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                column.width = max_length + 2  # Adicionar margem

        except Exception as e:
            self.log_updated.emit(f"FALHA NA FORMATAÇÃO FINAL: {str(e)}")
            raise

    def _clean_and_sort_output(self, sheet):
        """Método totalmente reformulado para evitar limites do Excel"""
        try:
            MAX_ROWS = 1048576  # Limite absoluto do Excel
            SAFE_LIMIT = MAX_ROWS - 1000  # Margem de segurança
            current_rows = sheet.max_row
            if current_rows <= 1:
                return  # Nada a processar

            all_rows = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row_idx >= SAFE_LIMIT:
                    self.log_updated.emit(f"Aviso: Dados truncados em {SAFE_LIMIT} linhas por segurança")
                    break
                if any(cell not in (None, "", 0) for cell in row):
                    all_rows.append(row)

            # Ordenação otimizada por Departamento (índice 3)
            all_rows.sort(key=lambda x: (str(x[3]).lower(), x[0]))

            # Limpeza radical e reinserção segura
            sheet.delete_rows(2, sheet.max_row)
            for row in all_rows:
                sheet.append(row)

            self.log_updated.emit(f"Processo quase finalizado,{len(all_rows)} linhas foram validadas, processando planilha a ser salva")
        except Exception as e:
            self.log_updated.emit(f"FALHA NA ORGANIZAÇÃO: {str(e)}")
            raise

    def _setup_styles(self):
        self.styles = {
            'date': NamedStyle(name="date", number_format='YYYY-MM-DD HH:MM:SS'),
            'minutes': NamedStyle(name="minutes", number_format='0.0'),
            'currency': NamedStyle(name="currency", number_format='R$ #,##0.00'),
            'duration': NamedStyle(name="duration", number_format='hh:mm:ss')
        }

    def _create_header(self, sheet):
        sheet.append(self.COLUNAS_SAIDA)
        for col, title in enumerate(self.COLUNAS_SAIDA, 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    def _get_output_path(self):
        base, _ = os.path.splitext(self._file_path)
        return f"{base}_leitura_agitel.xlsx"

    def _convert_date(self, value):
        if isinstance(value, dt_datetime):
            return xl_datetime.to_excel(value)
        return value

    def _convert_duration(self, value):
        if isinstance(value, dt_time):
            return value.hour / 24 + value.minute / 1440 + value.second / 86400
        elif isinstance(value, str):
            try:
                h, m, s = map(int, value.split(':'))
                return h / 24 + m / 1440 + s / 86400
            except:
                return 0.0
        return value

    def _duration_to_minutes(self, value):
        if isinstance(value, dt_time):
            return round(value.hour * 60 + value.minute + value.second / 60, 1)
        return 0.0

    def _parse_currency(self, value):
        try:
            return float(str(value).replace('R$', '').replace(',', '.').strip())
        except:
            return 0.0

    def _equalize_regiao(self, sheet):
        for row in sheet.iter_rows(min_row=2):
            cell = row[3]
            if cell.value:
                valor = str(cell.value).lower()
                if "fixo" in valor:
                    cell.value = "Fixo"
                elif any(x in valor for x in ["movel", "móvel"]):
                    cell.value = "Móvel"
                elif not valor.strip():
                    cell.value = "Intragrupo"

    def _find_header_row(self, sheet):
        essential_columns = {'data', 'origem', 'destino', 'duracao', 'preco'}
        for row in sheet.iter_rows(max_row=20):
            found = set()
            for cell in row:
                if cell.value:
                    normalized = self._normalize(str(cell.value))
                    if normalized in essential_columns:
                        found.add(normalized)
            if found >= essential_columns:
                return row
        return None

    def _get_column_indices(self, header_row):
        headers = [self._normalize(str(cell.value)) for cell in header_row]
        mapping = {
            'data': ['data', 'datachamada', 'datahora', 'datahorario'],
            'origem': ['origem', 'ramalorigem', 'setor', 'operador'],
            'servico': ['servico', 'tiposervico', 'serviço', 'tipochamada'],
            'regiao': ['regiao', 'região', 'localchamada', 'ddd'],
            'destino': ['destino', 'numerodestino', 'telefone', 'ramaldestino'],
            'duracao': ['duracao', 'tempochamada', 'tempogasto', 'duraçao'],
            'preco': ['preco', 'custochamada', 'valor', 'tarifa']
        }
        indices = {}
        for key, aliases in mapping.items():
            for alias in aliases:
                if self._normalize(alias) in headers:
                    indices[key] = headers.index(self._normalize(alias))
                    break
            else:
                raise ValueError(f"Coluna '{key}' não encontrada")
        return indices

    def _normalize(self, text):
        return ''.join(c for c in unicodedata.normalize('NFD', str(text).lower())
                       if not unicodedata.combining(c))

    def _handle_error(self, error):
        error_msg = f"Erro crítico: {str(error)}"
        self.error_occurred.emit(error_msg)
        logging.exception(error_msg)

    def _cleanup_resources(self, wb, output_wb):
        if wb:
            wb.close()
        if output_wb:
            output_wb.close()
        gc.collect()

    def _atualizar_progresso(self, index, progress_per_sheet):
        current_progress = int(index * progress_per_sheet)
        self.progress_updated.emit(min(current_progress, 100))