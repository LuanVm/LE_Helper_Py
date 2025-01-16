import os
import sys
import atexit
from PyQt6.QtWidgets import (
    QApplication, QLabel, QComboBox, QGraphicsOpacityEffect,
    QWidget, QGridLayout, QMessageBox, QTextEdit, QFileDialog, QLineEdit, QVBoxLayout, QPushButton, QHBoxLayout
)
from PyQt6.QtCore import QThreadPool, Qt, QSettings, QSize, QPropertyAnimation, QEasingCurve, Qt
from PyQt6.QtGui import QIcon, QTextCursor, QPixmap
from openpyxl import load_workbook

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'backend')))
from AutoBlume import Blume, AutomationTask, StopAutomation
from templates.GerenJanela import ResizableWindow
from templates.GerenEstilos import (
    estilo_sheet_light, estilo_label_light, estilo_combo_box_light, estilo_hover_light,
    campo_qline_light, estilo_log_light, estilo_sheet_dark, estilo_label_dark, estilo_combo_box_dark, estilo_hover_dark,
    campo_qline_dark, estilo_log_dark
)


class GuiAutoBlume(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.workbook = None
        self.sheet = None
        self.data_path = ""
        self.save_directory = ""
        self.threadpool = QThreadPool()
        self.stopflag = False
        self.init_ui()

    def init_ui(self):
        """Inicializa a interface gráfica."""
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(10, 10, 10, 10)
        self.layout.setSpacing(10)

        top_layout = QGridLayout()

        # Aplicar estilo ao QLabel
        self.label_salvamento = QLabel("Local de Salvamento:", self)
        self.label_salvamento.setStyleSheet(estilo_label_light())
        self.label_planilha = QLabel("Planilha de dados:", self)
        self.label_planilha.setStyleSheet(estilo_label_light())
        self.label_operadora = QLabel("Selecionar Operadora:", self)
        self.label_operadora.setStyleSheet(estilo_label_light())

        top_layout.addWidget(self.label_salvamento, 0, 0, alignment=Qt.AlignmentFlag.AlignLeft)
        top_layout.addWidget(self.label_planilha, 1, 0, alignment=Qt.AlignmentFlag.AlignLeft)
        top_layout.addWidget(self.label_operadora, 2, 0, alignment=Qt.AlignmentFlag.AlignLeft)

        # Diretório de Salvamento
        self.save_dir_field = QLineEdit(self)
        self.save_dir_field.setReadOnly(True)
        self.save_dir_field.setText(self.save_directory)
        self.save_dir_field.setStyleSheet(campo_qline_light())
        top_layout.addWidget(self.save_dir_field, 0, 1)

        self.save_dir_button = QPushButton("Selecionar Pasta", self)
        estilo_hover_light(self.save_dir_button)
        self.save_dir_button.clicked.connect(self.select_save_directory)
        top_layout.addWidget(self.save_dir_button, 0, 2)

        # Seleção de Planilha
        self.planilha_field = QLineEdit(self)
        self.planilha_field.setReadOnly(True)
        self.planilha_field.setText(self.data_path)
        self.planilha_field.setStyleSheet(campo_qline_light())
        top_layout.addWidget(self.planilha_field, 1, 1)

        self.planilha_button = QPushButton("Selecionar Planilha", self)
        estilo_hover_light(self.planilha_button)
        self.planilha_button.clicked.connect(self.select_data_file)
        top_layout.addWidget(self.planilha_button, 1, 2)

        # Seleção de Operadora
        self.operadora_combo = QComboBox(self)
        self.operadora_combo.addItem("Selecione uma planilha primeiro")
        self.operadora_combo.setStyleSheet(estilo_combo_box_light())
        top_layout.addWidget(self.operadora_combo, 2, 1)

        # Botão de Iniciar/Parar Automação
        self.confirm_button = QPushButton("Iniciar automação", self)
        estilo_hover_light(self.confirm_button)
        self.confirm_button.clicked.connect(self.toggle_automation)  # Alterado para toggle_automation
        top_layout.addWidget(self.confirm_button, 2, 2)

        self.layout.addLayout(top_layout)

        # Layout dos Logs
        logs_layout = QGridLayout()

        self.log_tecnico_area = QTextEdit(self)
        self.log_tecnico_area.setReadOnly(True)
        self.log_tecnico_area.setPlaceholderText("Log técnico")
        self.log_tecnico_area.setStyleSheet(estilo_log_light())
        logs_layout.addWidget(self.log_tecnico_area, 0, 0, 2, 2)

        self.faturas_coletadas_area = QTextEdit(self)
        self.faturas_coletadas_area.setReadOnly(True)
        self.faturas_coletadas_area.setPlaceholderText("Log faturas coletadas")
        self.faturas_coletadas_area.setStyleSheet(estilo_log_light())
        logs_layout.addWidget(self.faturas_coletadas_area, 0, 2, 2, 2)

        self.layout.addLayout(logs_layout)

    def select_data_file(self):
        last_dir = QSettings(self.parent.settings_path, QSettings.Format.IniFormat).value("last_open_dir", "")
        file_path, _ = QFileDialog.getOpenFileName(self, "Selecionar Planilha de Dados", last_dir, "Arquivos Excel (*.xlsx *.xlsm)")

        if file_path:
            if not os.path.isfile(file_path):
                QMessageBox.critical(self, "Erro", "O arquivo selecionado não é válido.")
            return
        self.load_data_file(file_path)

    def load_data_file(self, file_path):
        """Carrega a planilha de dados e atualiza a interface."""
        try:
            self.workbook = load_workbook(file_path)
            self.sheet = self.workbook.active
            self.data_path = file_path
            self.planilha_field.setText(file_path)

            # Atualiza o combo de operadoras
            self.operadora_combo.clear()
            operadoras = set()
            for row in self.sheet.iter_rows(min_row=2, values_only=True):
                if row[3]:  # Coluna 3: OPERADORA
                    operadoras.add(row[3])
            self.operadora_combo.addItems(operadoras)

            # Atualiza o caminho da planilha na instância de MainApp
            self.parent.data_path = file_path
            self.parent.save_settings()  # Salva as configurações
        except Exception as e:
            self.log_message(f"Erro: {str(e)}", area="tecnico")
            QMessageBox.critical(self, "Erro", f"Erro ao carregar a planilha: {e}")

    def select_save_directory(self):
        last_save_dir = QSettings(self.parent.settings_path, QSettings.Format.IniFormat).value("last_save_dir", "")
        dir_path = QFileDialog.getExistingDirectory(self, "Selecionar Diretório de Salvamento", last_save_dir)

        if dir_path and os.path.isdir(dir_path):
            self.save_directory = dir_path
            self.save_dir_field.setText(dir_path)
            
            # Atualiza o diretório de salvamento na instância de MainApp
            self.parent.save_directory = dir_path
            self.parent.save_settings()  # Salva as configurações
        else:
            self.save_directory = ""

    def toggle_automation(self):
        """Alterna entre iniciar e parar a automação."""
        if self.confirm_button.text() == "Iniciar automação":
            self.start_automation()
            self.confirm_button.setText("Parar automação")
            self.confirm_button.clicked.disconnect()  # Remove o conector anterior
            self.confirm_button.clicked.connect(self.toggle_automation)  # Reconecta ao mesmo método
        else:
            self.stop_automation()
            self.confirm_button.setText("Iniciar automação")
            self.confirm_button.clicked.disconnect()  # Remove o conector anterior
            self.confirm_button.clicked.connect(self.toggle_automation)  # Reconecta ao mesmo método
        
    def start_automation(self):
        """Inicia o processo de automação."""
        selected_operadora = self.operadora_combo.currentText()

        if not self.save_directory or not os.path.isdir(self.save_directory):
            QMessageBox.warning(self, "Erro", "Selecione um diretório de salvamento primeiro.")
            return

        if not selected_operadora:
            QMessageBox.warning(self, "Erro", "Selecione uma operadora.")
            return

        if not self.data_path or not os.path.isfile(self.data_path):
            QMessageBox.warning(self, "Erro", "Selecione uma planilha de dados primeiro.")
            return

        if selected_operadora.upper() == "BLUME":
            try:
                self.automator = Blume(self, self.data_path)  # Salva o automator como atributo da classe
            except Exception as e:
                self.log_message(f"Erro: {str(e)}", area="tecnico")
                QMessageBox.critical(self, "Erro", f"Erro ao criar o automator Blume: {e}")
                return
        else:
            QMessageBox.warning(self, "Erro", f"Automação para a operadora {selected_operadora} não está implementada.")
            return

        try:
            # Cria a tarefa de automação
            task = AutomationTask(self.automator, self.get_user_data(selected_operadora), self.log_message)
            
            # Inicia a tarefa no thread pool
            self.threadpool.start(task)
        except Exception as e:
            self.log_message(f"Erro: {str(e)}", area="tecnico")
            QMessageBox.critical(self, "Erro", f"Erro ao iniciar a tarefa de automação: {e}")

    def stop_automation(self):
        """Para a automação em execução."""
        if hasattr(self, 'automator'):
            stop_automation = StopAutomation(self.automator)
            stop_automation.stop()
            self.confirm_button.setText("Iniciar automação")  # Garante que o botão volte ao estado inicial
        else:
            QMessageBox.warning(self, "Erro", "Nenhuma automação em execução para interromper.")

    def get_user_data(self, selected_operadora):
        """Filtra os dados da operadora selecionada."""
        user_data = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[3] == selected_operadora and row[11] != 'COLETADO IA':  # Coluna 3: OPERADORA, Coluna 11: STATUS
                user_data.append({
                    "FORNECEDOR": row[0],
                    "REFERÊNCIA": row[1],
                    "CLIENTE": row[2],
                    "OPERADORA": row[3],
                    "IDENTIFICAÇÃO": row[4],
                    "CÓDIGO": row[5],
                    "PA": row[6],
                    "INDENTIFICAÇÃO INTERNA": row[7],
                    "LOGIN": row[8],
                    "SENHA": row[9],
                    "VENCIMENTO": row[10],
                    "STATUS": row[11],
                    "NOMENCLATURA": row[12]
                })
        return user_data

    def log_message(self, message, area="tecnico", color=None):
        """Exibe mensagens nos logs adequados, com suporte para cores."""
        if area == "tecnico":
            log_area = self.log_tecnico_area
        elif area == "faturas":
            log_area = self.faturas_coletadas_area
        else:
            return  # Área de log inválida

        # Aplica a cor ao texto, se especificada
        if color:
            message = f'<span style="color: {color};">{message}</span>'

        # Adiciona a mensagem ao log
        log_area.append(message)

        # Move o cursor para o final do log
        cursor = log_area.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)
        log_area.setTextCursor(cursor)


class MainApp(ResizableWindow):
    def __init__(self):
        super().__init__()

        # Configuração da interface gráfica
        self.setStyleSheet(estilo_sheet_light())

        # Remove a barra de título padrão do sistema operacional
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)  # Fundo transparente

        # Configurações persistentes
        self.settings_path = "config.ini"
        self.save_directory = ""
        self.data_path = ""

        # Estado inicial do modo (light mode)
        self.modo_escuro = False  # Inicializa o atributo antes de carregar as configurações

        # Configurar a janela principal
        self.setWindowTitle("LE Helper")
        self.setGeometry(100, 100, 1100, 650)

        # Cria um widget central com bordas arredondadas
        self.central_widget = QWidget(self)
        self.central_widget.setObjectName("central_widget")
        self.central_widget.setStyleSheet(estilo_sheet_light())
        self.setCentralWidget(self.central_widget)

        # Layout principal
        self.layout = QVBoxLayout(self.central_widget)
        self.layout.setContentsMargins(10, 10, 10, 10)
        self.layout.setSpacing(10)

        # Cria uma barra de título personalizada
        self.barra_titulo = QWidget(self.central_widget)
        self.barra_titulo.setObjectName("barra_titulo")
        self.barra_titulo.setFixedHeight(30)

        layout_titulo = QHBoxLayout(self.barra_titulo)
        layout_titulo.setContentsMargins(5, 0, 5, 0)
        layout_titulo.setSpacing(5)

        # Ícone do título
        caminho_base = os.path.join(os.path.dirname(__file__), "..", "frontend", "static", "icons")
        icone_titulo = QLabel()
        caminho_icone = os.path.join(caminho_base, "logo.png")
        if os.path.exists(caminho_icone):
            icone_titulo.setPixmap(QPixmap(caminho_icone).scaled(
                24, 24, 
                Qt.AspectRatioMode.IgnoreAspectRatio,
                Qt.TransformationMode.SmoothTransformation
            ))
        layout_titulo.addWidget(icone_titulo)

        # ComboBox para selecionar funcionalidades
        self.funcionalidades_combo = QComboBox(self.barra_titulo)
        self.funcionalidades_combo.addItem("Automação coleta")
        self.funcionalidades_combo.setStyleSheet(estilo_combo_box_light())
        self.funcionalidades_combo.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.funcionalidades_combo.currentTextChanged.connect(self.mudar_funcionalidade)
        layout_titulo.addWidget(self.funcionalidades_combo)

        # Espaçador para alinhar os botões à direita
        layout_titulo.addStretch()

        # Botão de alternância de modo (dark/light)
        self.botao_modo = QPushButton(self.barra_titulo)
        self.botao_modo.setObjectName("botao_modo")
        self.botao_modo.setIcon(QIcon("frontend/static/icons/ui_light.png"))
        self.botao_modo.setFixedSize(QSize(20, 20))
        self.botao_modo.setStyleSheet("background-color: transparent; border: none;")  # Estilo transparente
        self.botao_modo.clicked.connect(self.alternar_modo)
        layout_titulo.addWidget(self.botao_modo)

        # Botão de minimizar
        self.botao_minimizar = QPushButton(self.barra_titulo)
        self.botao_minimizar.setObjectName("botao_minimizar")
        self.botao_minimizar.setIcon(QIcon("frontend/static/icons/ui_minimize_light.png"))
        self.botao_minimizar.setFixedSize(QSize(20, 20))
        self.botao_minimizar.setStyleSheet("background-color: transparent; border: none;")  # Estilo transparente
        self.botao_minimizar.clicked.connect(self.showMinimized)
        layout_titulo.addWidget(self.botao_minimizar)

        # Botão de fechar
        self.botao_fechar = QPushButton(self.barra_titulo)
        self.botao_fechar.setObjectName("botao_fechar")
        self.botao_fechar.setIcon(QIcon("frontend/static/icons/ui_exit_light.png"))
        self.botao_fechar.setFixedSize(QSize(20, 20))
        self.botao_fechar.setStyleSheet("background-color: transparent; border: none;")  # Estilo transparente
        self.botao_fechar.clicked.connect(self.close)
        layout_titulo.addWidget(self.botao_fechar)

        # Adiciona a barra de título personalizada ao layout principal
        self.layout.addWidget(self.barra_titulo)

        # Widget central e layout principal
        self.central_content = QWidget(self.central_widget)
        self.layout.addWidget(self.central_content)
        self.content_layout = QVBoxLayout(self.central_content)

        # Inicializa a funcionalidade padrão
        self.gui_auto_blume = GuiAutoBlume(self)
        self.content_layout.addWidget(self.gui_auto_blume)

        # Carrega as configurações salvas
        self.load_settings()  # Agora modo_escuro já está inicializado

        # Aplica o estilo inicial
        if self.modo_escuro:
            self.aplicar_estilo_dark()
        else:
            self.aplicar_estilo_light()

    def closeEvent(self, event):
        """Sobrescreve o método closeEvent para encerrar a aplicação de forma limpa."""
        # Verifica se o automator existe antes de tentar acessá-lo
        if hasattr(self.gui_auto_blume, 'automator'):
            # Encerra a automação, se estiver em execução
            stop_automation = StopAutomation(self.gui_auto_blume.automator)
            stop_automation.stop()

            # Fecha todos os navegadores abertos
            if hasattr(self.gui_auto_blume.automator, 'drivers'):
                for driver in self.gui_auto_blume.automator.drivers:
                    try:
                        driver.quit()
                    except Exception as e:
                        print(f"Erro ao fechar o navegador: {e}")

        # Fecha a aplicação
        event.accept()  # Aceita o evento de fechamento

    def close_application(self):
        """Encerra a aplicação de forma assíncrona."""
        self.close()

    def load_settings(self):
        """Carrega as configurações salvas no arquivo de configurações."""
        settings = QSettings(self.settings_path, QSettings.Format.IniFormat)
        
        # Carrega o diretório de salvamento
        self.save_directory = settings.value("save_directory", "")
        self.gui_auto_blume.save_directory = self.save_directory
        self.gui_auto_blume.save_dir_field.setText(self.save_directory)
        
        # Carrega o caminho da planilha
        self.data_path = settings.value("data_path", "")
        self.gui_auto_blume.data_path = self.data_path
        self.gui_auto_blume.planilha_field.setText(self.data_path)
        
        # Carrega o último diretório aberto (para uso futuro)
        if self.data_path:
            self.gui_auto_blume.load_data_file(self.data_path)
        
        # Carrega o tema salvo
        self.modo_escuro = settings.value("dark_mode", False, type=bool)
        if self.modo_escuro:
            self.aplicar_estilo_dark()
        else:
            self.aplicar_estilo_light()

    def save_settings(self):
        """Salva as configurações atuais no arquivo de configurações."""
        settings = QSettings(self.settings_path, QSettings.Format.IniFormat)
        
        # Salva o diretório de salvamento
        settings.setValue("save_directory", self.save_directory)
        
        # Salva o caminho da planilha
        settings.setValue("data_path", self.data_path)
        
        # Salva o último diretório aberto (para uso futuro)
        if self.data_path:
            settings.setValue("last_open_dir", os.path.dirname(self.data_path))
        
        # Salva o último diretório de salvamento (para uso futuro)
        if self.save_directory:
            settings.setValue("last_save_dir", self.save_directory)
        
        # Salva o tema atual
        settings.setValue("dark_mode", self.modo_escuro)

    def alternar_modo(self):
        """Alterna entre os modos light e dark com uma transição suave."""
        if self.modo_escuro:
            # Aplica o estilo light imediatamente
            self.aplicar_estilo_light()
        else:
            # Aplica o estilo dark imediatamente
            self.aplicar_estilo_dark()

        # Inverte o estado do modo
        self.modo_escuro = not self.modo_escuro

        self.botao_modo.setIcon(QIcon("frontend/static/icons/ui_dark.png" if self.modo_escuro else "frontend/static/icons/ui_light.png"))
        self.botao_minimizar.setIcon(QIcon("frontend/static/icons/ui_minimize_dark.png" if self.modo_escuro else "frontend/static/icons/ui_minimize_light.png"))
        self.botao_fechar.setIcon(QIcon("frontend/static/icons/ui_exit_dark.png" if self.modo_escuro else "frontend/static/icons/ui_exit_light.png"))

        # Anima o efeito de fade
        self.animar_fade()

    def animar_fade(self):
        """Anima o efeito de fade em um widget."""
        # Cria um efeito de opacidade
        opacity_effect = QGraphicsOpacityEffect(self.central_widget)
        self.central_widget.setGraphicsEffect(opacity_effect)

        # Cria a animação de fade
        animacao = QPropertyAnimation(opacity_effect, b"opacity")
        animacao.setDuration(300)  # Duração da animação em milissegundos
        animacao.setEasingCurve(QEasingCurve.Type.InOutQuad)  # Curva de animação suave
        animacao.setStartValue(1.0)
        animacao.setEndValue(0.5)  # Reduz a opacidade para 0.5

        # Cria uma segunda animação para retornar à opacidade total
        animacao_reversa = QPropertyAnimation(opacity_effect, b"opacity")
        animacao_reversa.setDuration(300)
        animacao_reversa.setEasingCurve(QEasingCurve.Type.InOutQuad)
        animacao_reversa.setStartValue(0.5)
        animacao_reversa.setEndValue(1.0)

        # Remove o efeito após a animação reversa
        animacao_reversa.finished.connect(lambda: self.central_widget.setGraphicsEffect(None))

        # Conecta as animações
        animacao.finished.connect(animacao_reversa.start)

        # Inicia a animação
        animacao.start()

    def aplicar_estilo_light(self):
        """Aplica o estilo light mode a todos os componentes."""
        self.setStyleSheet(estilo_sheet_light())
        self.central_widget.setStyleSheet(estilo_sheet_light())
        self.barra_titulo.setStyleSheet(estilo_sheet_light())
        self.funcionalidades_combo.setStyleSheet(estilo_combo_box_light())
        self.gui_auto_blume.setStyleSheet(estilo_sheet_light())
        self.gui_auto_blume.save_dir_field.setPlaceholderText('Selecione o diretório de salvamento...')
        self.gui_auto_blume.save_dir_field.setStyleSheet(campo_qline_light())
        self.gui_auto_blume.planilha_field.setPlaceholderText('Selecione a planilha de dados...')
        self.gui_auto_blume.planilha_field.setStyleSheet(campo_qline_light())
        self.gui_auto_blume.operadora_combo.setStyleSheet(estilo_combo_box_light())
        self.gui_auto_blume.log_tecnico_area.setStyleSheet(estilo_log_light())
        self.gui_auto_blume.faturas_coletadas_area.setStyleSheet(estilo_log_light())
        estilo_hover_light(self.gui_auto_blume.save_dir_button)
        estilo_hover_light(self.gui_auto_blume.planilha_button)
        estilo_hover_light(self.gui_auto_blume.confirm_button)
        
        # Aplicar estilo aos QLabel
        self.gui_auto_blume.label_salvamento.setStyleSheet(estilo_label_light())
        self.gui_auto_blume.label_planilha.setStyleSheet(estilo_label_light())
        self.gui_auto_blume.label_operadora.setStyleSheet(estilo_label_light())

        # Estilo dos botões da barra de título
        self.botao_modo.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
                border-radius: 5px;
            }
        """)
        self.botao_minimizar.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
                border-radius: 5px;
            }
        """)
        self.botao_fechar.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
            }
            QPushButton:hover {
                background-color: #ff6b6b;
                border-radius: 5px;
            }
        """)

    def aplicar_estilo_dark(self):
        """Aplica o estilo dark mode a todos os componentes."""
        self.setStyleSheet(estilo_sheet_dark())
        self.central_widget.setStyleSheet(estilo_sheet_dark())
        self.barra_titulo.setStyleSheet(estilo_sheet_dark())
        self.funcionalidades_combo.setStyleSheet(estilo_combo_box_dark())
        self.gui_auto_blume.setStyleSheet(estilo_sheet_dark())
        self.gui_auto_blume.save_dir_field.setStyleSheet(campo_qline_dark())
        self.gui_auto_blume.planilha_field.setStyleSheet(campo_qline_dark())
        self.gui_auto_blume.operadora_combo.setStyleSheet(estilo_combo_box_dark())
        self.gui_auto_blume.log_tecnico_area.setStyleSheet(estilo_log_dark())
        self.gui_auto_blume.faturas_coletadas_area.setStyleSheet(estilo_log_dark())
        estilo_hover_dark(self.gui_auto_blume.save_dir_button)
        estilo_hover_dark(self.gui_auto_blume.planilha_button)
        estilo_hover_dark(self.gui_auto_blume.confirm_button)
        self.gui_auto_blume.label_salvamento.setStyleSheet(estilo_label_dark())
        self.gui_auto_blume.label_planilha.setStyleSheet(estilo_label_dark())
        self.gui_auto_blume.label_operadora.setStyleSheet(estilo_label_dark())

        # Estilo dos botões da barra de título
        self.botao_modo.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
            }
            QPushButton:hover {
                background-color: #444444;
                border-radius: 5px;
            }
        """)
        self.botao_minimizar.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
            }
            QPushButton:hover {
                background-color: #444444;
                border-radius: 5px;
            }
        """)
        self.botao_fechar.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
            }
            QPushButton:hover {
                background-color: #ff6b6b;
                border-radius: 5px;
            }
        """)

    def mudar_funcionalidade(self, funcionalidade):
        """Alterna entre as funcionalidades disponíveis."""
        if funcionalidade == "Automação coleta":
            self.content_layout.addWidget(self.gui_auto_blume)
        else:
            # Adicione outras funcionalidades aqui
            pass


if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_window = MainApp()
    main_window.show()
    sys.exit(app.exec())