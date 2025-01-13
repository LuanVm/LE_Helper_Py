import os
import time
import shutil
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from PyQt6.QtCore import QRunnable, pyqtSlot, QMutex
from PyPDF2 import PdfReader
from openpyxl import load_workbook


class AutomationTask(QRunnable):
    def __init__(self, automator, user_data, log_function):
        super().__init__()
        self.automator = automator
        self.user_data = user_data
        self.log_function = log_function

    @pyqtSlot()
    def run(self):
        try:
            self.log_function("Iniciando automação para o usuário...")
            self.automator.run_automation(self.user_data)
            self.log_function("Automação concluída com sucesso.")
        except Exception as e:
            self.log_function(f"Erro durante a automação: {str(e)}")


class Blume:
    def __init__(self, parent, data_path):
        self.parent = parent
        self.data_path = data_path
        self.workbook = load_workbook(data_path)
        self.sheet = self.workbook.active
        self.mutex = QMutex()

    def initialize_browser(self):
        try:
            options = webdriver.ChromeOptions()
            options.add_argument("--disable-extensions")
            options.add_argument("--disable-popup-blocking")

            self.parent.log_message("Inicializando o navegador...")
            driver = webdriver.Chrome(
                service=Service(ChromeDriverManager().install()),
                options=options
            )
            self.parent.log_message("Navegador inicializado com sucesso.")
            return driver
        except Exception as e:
            self.parent.log_message(f"Erro ao iniciar o navegador: {type(e).__name__} - {e}")
            raise

    def run_automation(self, user_data):
        self.parent.log_message("Iniciando processo de automação para a operadora Blume...", area="tecnico")

        if self.verificar_coleta_finalizada():
            self.parent.log_message("Nenhuma coleta pendente. Finalizando execução.")
            return

        for user in user_data:
            if user['STATUS'] == 'COLETADO IA' or user['STATUS'] == 'INDISPONIVEL':
                self.parent.log_message(f"Estrutura {user['LOGIN']} já processada ou indisponível, ignorando...", area="tecnico")
                continue

            driver = None
            try:
                driver = self.initialize_browser()
                wait = WebDriverWait(driver, 15)
                driver.get("https://portal.blumetelecom.com.br")
                time.sleep(2)
                self.parent.log_message("Página de login carregada, iniciando login...", area="tecnico")
                self.login(driver, wait, user)

                time.sleep(2)
                driver.get("https://portal.blumetelecom.com.br/billings")

                self.parent.log_message("Login bem-sucedido, iniciando processamento dos boletos...", area="tecnico")
                self.process_boletos(driver, wait, user)

            except Exception as e:
                self.parent.log_message(f"Erro durante o processamento do usuário {user['LOGIN']}: {str(e)}", area="tecnico")
            finally:
                if driver:
                    self.parent.log_message("Fechando o navegador...", area="tecnico")
                    driver.quit()

        self.parent.log_message("Todos os boletos foram processados.", area="tecnico")

    def login(self, driver, wait, user_data):
        try:
            self.parent.log_message("Tentando acessar a página de login...")

            # Espera até que a página esteja completamente carregada
            wait.until(lambda driver: driver.execute_script("return document.readyState") == "complete")

            self.parent.log_message("Procurando campo de email...")
            email_field = wait.until(EC.presence_of_element_located((By.ID, "loginUsername")))
            self.parent.log_message("Campo de email encontrado.")

            # Rola a página até o campo de e-mail (caso ele não esteja visível)
            driver.execute_script("arguments[0].scrollIntoView(true);", email_field)
            time.sleep(1)  # Pequena pausa para garantir que o campo esteja visível

            # Tenta preencher o campo de e-mail sem limpar
            try:
                self.parent.log_message("Preenchendo campo de email...")
                email_field.send_keys(user_data['LOGIN'])
            except Exception as e:
                self.parent.log_message(f"Erro ao preencher o campo de e-mail: {e}")
                raise

            self.parent.log_message("Procurando campo de senha...")
            password_field = wait.until(EC.element_to_be_clickable((By.NAME, "password")))
            self.parent.log_message("Campo de senha encontrado.")

            # Rola a página até o campo de senha (caso ele não esteja visível)
            driver.execute_script("arguments[0].scrollIntoView(true);", password_field)

            self.parent.log_message("Preenchendo campo de senha...")
            password_field.clear()
            password_field.send_keys(user_data['SENHA'])

            self.parent.log_message("Procurando botão de login...")
            login_button = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "MuiButton-label")))
            self.parent.log_message("Botão de login encontrado.")

            # Rola a página até o botão de login (caso ele não esteja visível)
            driver.execute_script("arguments[0].scrollIntoView(true);", login_button)

            self.parent.log_message("Clicando no botão de login...")
            login_button.click()

            self.parent.log_message("Login realizado com sucesso.")

        except Exception as e:
            self.parent.log_message(f"Erro durante o login: {e}")
            raise

    def process_boletos(self, driver, wait, user_data):
        try:
            processed_boleto_ids = set()

            while True:
                self.parent.log_message("Acessando página de billings...")
                driver.get("https://portal.blumetelecom.com.br/billings")

                # Verifica a presença do texto "Você não possui faturas em aberto"
                try:
                    no_invoices_message = wait.until(
                        EC.presence_of_element_located((By.XPATH, "//h5[contains(text(), 'Você não possui faturas em aberto')]"))
                    )
                    self.parent.log_message("Mensagem 'Você não possui faturas em aberto' encontrada.")
                    
                    # Atualiza o status para 'INDISPONIVEL' e encerra o processamento para este cliente
                    self.atualizar_status_na_planilha(user_data['LOGIN'], 'INDISPONIVEL')
                    self.parent.log_message(f"Status atualizado para 'INDISPONIVEL' para o cliente {user_data['LOGIN']}.")
                    return  # Encerra o método process_boletos para este cliente
                except Exception as e:
                    self.parent.log_message(f"Não encontrou a mensagem 'Você não possui faturas em aberto': {e}")
                    # Caso não encontre a mensagem, prossegue com a coleta de boletos
                    pass

                # Captura os botões 'Pagar boleto'
                try:
                    boleto_buttons = wait.until(
                        EC.presence_of_all_elements_located((By.XPATH, "//span[text()='Pagar boleto']"))
                    )
                    self.parent.log_message(f"Botões 'Pagar boleto' encontrados: {len(boleto_buttons)}")
                except Exception as e:
                    self.parent.log_message(f"Erro ao localizar botões 'Pagar boleto': {e}")
                    break

                for index, button in enumerate(boleto_buttons):
                    boleto_id = f"button-{index}"
                    if boleto_id in processed_boleto_ids:
                        continue

                    self.parent.log_message(f"Preparando para processar 'Pagar boleto - {index + 1}'...")

                    # Certifique-se de que o botão é clicável
                    try:
                        wait.until(EC.element_to_be_clickable(button))
                        driver.execute_script("arguments[0].scrollIntoView(true);", button)
                        button.click()
                    except Exception as e:
                        self.parent.log_message(f"Erro ao clicar no botão 'Pagar boleto - {index + 1}': {e}")
                        continue

                    processed_boleto_ids.add(boleto_id)

                    # Aguarda o redirecionamento para a página de detalhes do boleto
                    try:
                        wait.until(EC.url_contains("billings"))
                        self.download_boleto(wait, user_data, index + 1)
                        time.sleep(1)
                    except Exception as e:
                        self.parent.log_message(f"Erro após clicar em 'Pagar boleto - {index + 1}': {e}")
                        continue

                # Verifica se ainda há boletos não processados
                boletos_disponiveis = [
                    button for idx, button in enumerate(boleto_buttons)
                    if f"button-{idx}" not in processed_boleto_ids
                ]
                if not boletos_disponiveis:
                    self.parent.log_message("Todos os boletos processados ou nenhum novo encontrado.")
                    break

        except Exception as e:
            self.parent.log_message(f"Erro geral ao processar boletos: {e}")
            raise

    def download_boleto(self, wait, user_data, index):
        try:
            download_button = wait.until(
                EC.presence_of_element_located((By.XPATH, "//p[text()='Baixar boleto']"))
            )
            download_button.click()

            self.parent.log_message("Aguardando conclusão do download...")
            self.wait_for_download()

            download_dir = os.path.join(os.path.expanduser('~'), 'Downloads')
            latest_file = self.get_latest_file(download_dir)

            if latest_file:
                self.parent.log_message(f"Boleto baixado: {latest_file}", area="tecnico")
                contrato = self.extrair_contrato_pdf(latest_file)
                if contrato:  # Validar antes de continuar
                    self.handle_downloaded_file(contrato, latest_file, user_data)
            else:
                self.parent.log_message(f"Erro ao baixar o boleto {index}: Arquivo não encontrado.")

        except Exception as e:
            self.parent.log_message(f"Erro ao processar o download do boleto {index}: {e}")

    def wait_for_download(self):
        download_dir = os.path.join(os.path.expanduser('~'), 'Downloads')
        timeout = 60
        while timeout > 0:
            pdf_files = [fname for fname in os.listdir(download_dir) if fname.endswith(".pdf")]
            if pdf_files:
                latest_file = os.path.join(download_dir, max(pdf_files, key=lambda f: os.path.getmtime(os.path.join(download_dir, f))))
                if not latest_file.endswith(".crdownload") and os.path.getsize(latest_file) > 0:
                    time.sleep(2)
                    return
            time.sleep(1)
            timeout -= 1
        raise Exception("Tempo de download excedido ou arquivo incompleto.")

    def extrair_contrato_pdf(self, pdf_path):
        try:
            if not os.path.exists(pdf_path) or os.path.getsize(pdf_path) == 0:
                self.parent.log_message(f"Arquivo PDF inválido: {pdf_path}")
                return None
            leitor = PdfReader(pdf_path)
            texto = ""
            for pagina in leitor.pages:
                texto += pagina.extract_text()

            # Regex ajustado para capturar especificamente o contrato no formato correto
            padrao = r'\b0{3,}(\d{3,})\b'
            match = re.search(padrao, texto)

            if match:
                contrato = match.group(1)  # Pega o grupo capturado sem os zeros à esquerda
                return contrato
            else:
                return None
        except Exception as e:
            self.parent.log_message(f"Erro ao processar o PDF {pdf_path}: {str(e)}")
            return None

    def get_latest_file(self, directory):
        try:
            files = [os.path.join(directory, fname) for fname in os.listdir(directory) if fname.endswith('.pdf')]
            return max(files, key=os.path.getmtime) if files else None
        except Exception as e:
            self.parent.log_message(f"Erro ao obter o arquivo mais recente: {e}")
            return None

    def handle_downloaded_file(self, contrato, latest_file, user_data):
        if contrato:
            contrato = contrato.lstrip('0')
            self.parent.log_message(f"Número do contrato extraído: {contrato}")

            # Verifica se o contrato existe na planilha
            if self.verificar_contrato_na_planilha(contrato):
                nomenclatura = self.obter_nomenclatura_por_contrato(contrato)
                destino = os.path.join(self.parent.save_directory, f"{nomenclatura}.pdf")
                shutil.move(latest_file, destino)
                self.parent.log_message(f"{nomenclatura}", area="faturas")

                # Atualiza o status na planilha
                self.atualizar_status_na_planilha(contrato, 'COLETADO IA')
            else:
                self.parent.log_message(f"Contrato {contrato} não encontrado na planilha.")
                self.mover_boleto_nao_encontrado(latest_file)
        else:
            self.parent.log_message(f"Contrato não extraído do PDF: {latest_file}")
            self.mover_boleto_nao_encontrado(latest_file)

    def mover_boleto_nao_encontrado(self, pdf_path):
        output_dir = os.path.join(self.parent.save_directory, "Boleto não encontrado")
        os.makedirs(output_dir, exist_ok=True)

        base_filename = "Boleto_nao_encontrado"
        file_index = 1
        new_filename = f"{base_filename}.pdf"
        while os.path.exists(os.path.join(output_dir, new_filename)):
            new_filename = f"{base_filename}_{file_index}.pdf"
            file_index += 1

        destino = os.path.join(output_dir, new_filename)
        shutil.move(pdf_path, destino)
        self.parent.log_message(f"Boleto movido para 'Boleto não encontrado' como: {destino}")

    def verificar_contrato_na_planilha(self, contrato):
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if str(row[4]).lstrip('0') == contrato:  # Coluna 4: IDENTIFICAÇÃO
                return True
        return False

    def obter_nomenclatura_por_contrato(self, contrato):
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if str(row[4]).lstrip('0') == contrato:  # Coluna 4: IDENTIFICAÇÃO
                return row[12]  # Coluna 12: NOMENCLATURA
        return None

    def atualizar_status_na_planilha(self, login, status):
        self.mutex.lock()
        try:
            for row in self.sheet.iter_rows(min_row=2):
                if str(row[8].value) == login:  # Coluna 8: LOGIN
                    row[11].value = status  # Coluna 11: STATUS
                    self.workbook.save(self.data_path)
                    self.parent.log_message(f"Status atualizado para '{status}' no login {login}.")
                    break
        except Exception as e:
            self.parent.log_message(f"Erro ao atualizar status na planilha: {e}")
        finally:
            self.mutex.unlock()

    def verificar_coleta_finalizada(self):
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[11] != 'COLETADO IA':  # Coluna 11: STATUS
                return False
        return True